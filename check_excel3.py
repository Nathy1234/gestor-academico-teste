import openpyxl
wb = openpyxl.load_workbook('CURSOS INOVA - LINKS (1).xlsx')

# Checar TODAS as linhas do profissionalizantes que tem dado
ws = wb['PROFISSIONALIZANTES INOVA']
print(f'=== PROFISSIONALIZANTES INOVA - linhas nao vazias apos 43 (total={ws.max_row}) ===')
count = 0
for i, row in enumerate(ws.iter_rows(min_row=44, values_only=True), 44):
    vals = [str(c or '')[:60] for c in row[:9]]
    if any(v.strip() for v in vals):
        print(f'L{i}: {vals}')
        count += 1
        if count >= 60:
            print('...')
            break

if count == 0:
    print('NENHUMA LINHA COM DADOS ENCONTRADA')

print()
# Checar POS MATRIZES - ver quantas matrizes tem
ws2 = wb['PÓS- MATRIZES INOVA']
print('=== POS MATRIZES: procurando headers de curso ===')
cursos_matrix = []
for i, row in enumerate(ws2.iter_rows(min_row=48, values_only=True), 48):
    c0 = str(row[0] or '').strip()
    c1 = str(row[1] or '').strip()
    c2 = str(row[2] or '').strip() if len(row) > 2 else ''
    # Nome do curso: col B com valor, col A vazio ou col C vazio
    if c1 and not c2 and not c0:
        cursos_matrix.append((i, c1[:70]))
    elif c1 and not c2 and c0:
        cursos_matrix.append((i, f'[{c0}] {c1[:70]}'))

print(f'Total de headers de curso encontrados: {len(cursos_matrix)}')
for linha, nome in cursos_matrix:
    print(f'  L{linha}: {nome}')

print()
# Ver profissionalizante com matrizes em outro sheet - TITULACOES POS
for name in wb.sheetnames:
    if 'TITULA' in name.upper():
        ws3 = wb[name]
        print(f'=== {name}: {ws3.max_row} linhas - buscando prof data ===')
        count = 0
        for i, row in enumerate(ws3.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            print(f'L{i}: {[str(c or "")[:50] for c in row[:5]]}')
        break
