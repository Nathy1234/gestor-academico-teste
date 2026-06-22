import openpyxl
wb = openpyxl.load_workbook('CURSOS INOVA - LINKS (1).xlsx')

# Profissionalizantes
ws = wb['PROFISSIONALIZANTES INOVA']
print('=== PROFISSIONALIZANTES - primeiras 50 linhas ===')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
    vals = [str(c or '')[:50] for c in row[:9]]
    if any(v.strip() for v in vals):
        print(f'L{i}: {vals}')

print()
print('=== POS MATRIZES - estrutura ===')
ws2 = wb['PÓS- MATRIZES INOVA']
print(f'Total linhas: {ws2.max_row}')
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f'L{i}: {[str(c or "")[:40] for c in row[:10]]}')

print()
print('=== TITULACOES POS - primeiras 10 linhas ===')
# find the sheet
for name in wb.sheetnames:
    if 'TITULA' in name.upper():
        ws3 = wb[name]
        print(f'Aba: {name}, {ws3.max_row} linhas')
        for i, row in enumerate(ws3.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            print(f'L{i}: {[str(c or "")[:45] for c in row[:5]]}')
        break
