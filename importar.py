import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from app import app, db, User, Course, Discipline, Coupon, Refund
from datetime import datetime
import openpyxl

def is_numero(val):
    """Verifica se val é um número de linha (1, 2, 3...)."""
    try:
        return int(str(val).strip()) > 0
    except:
        return False

def limpar_horas(val):
    """Retorna apenas a parte numérica das horas, ou vazio se não for horas."""
    if val is None:
        return ''
    s = str(val).strip()
    # Extrai o primeiro número encontrado (ex: "10/NÃO FOI..." -> "10", "360h" -> "360")
    m = re.match(r'^(\d+\.?\d*)', s)
    if m:
        return m.group(1)
    return ''

def limpar_valor(val):
    """Retorna o valor como string, ou vazio se for texto não-numérico."""
    if val is None:
        return ''
    s = str(val).strip()
    if s in ('-', 'GRATUITO', 'gratuito', 'Gratuito', ''):
        return s if s in ('GRATUITO', 'Gratuito', 'gratuito') else ''
    return s

def pd(v):
    if isinstance(v, datetime):
        return v.date()
    return None

def add_course(admin_id, nome, tipo, area, horas, valor, link, link_img, desc, obs, status, insersor, meses='', cupom='', dono=''):
    if not nome or str(nome).strip() == '':
        return None
    c = Course(
        nome=str(nome).strip(),
        tipo=tipo,
        area=str(area or '').strip()[:100],
        horas=str(horas or '').strip()[:20],
        meses=str(meses or '').strip()[:20],
        valor=str(valor or '').strip()[:50],
        link_venda=str(link or '').strip(),
        link_imagem=str(link_img or '').strip(),
        descricao=str(desc or '').strip(),
        obs=str(obs or '').strip(),
        status=status,
        insersor=str(insersor or '').strip()[:100],
        cupom=str(cupom or '').strip()[:100],
        dono=str(dono or '').strip(),
        created_by=admin_id
    )
    db.session.add(c)
    return c

with app.app_context():
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        print("ERRO: usuario admin nao encontrado.")
        sys.exit(1)

    print("Limpando dados antigos...")
    Discipline.query.delete()
    Course.query.delete()
    Coupon.query.delete()
    Refund.query.delete()
    db.session.commit()

    caminho = os.path.join(os.path.dirname(__file__), 'CURSOS INOVA - LINKS (1).xlsx')
    wb = openpyxl.load_workbook(caminho)
    total = 0

    # ── PÓS-GRADUAÇÃO (PÓS- MATRIZES INOVA) ─────────────────────────────
    # Estrutura: linha 1=aviso, linha 2=cabecalho, linha 3+=dados
    # Cabecalho: Nº | PÓS INOVA | ÁREA | HORAS | MESES | VALOR | INSERSOR | LINK VENDA | OBS
    # ATENÇÃO: a aba tem matrizes de disciplinas embutidas abaixo dos cursos.
    # Só importar linhas onde col[0] é um número inteiro (Nº do curso).
    for shname in wb.sheetnames:
        if 'MATRIZES' in shname.upper():
            print(f"\nImportando Pos-Graduacao ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not is_numero(row[0]):
                    continue  # pula linhas de matriz de disciplinas e subtítulos
                # Linha de disciplina da matriz: col[2] = "38h" (CH), não área
                area_val = str(row[2] or '').strip()
                if area_val.endswith('h') and area_val[:-1].isdigit():
                    continue
                nome  = row[1]
                area  = row[2]
                horas = limpar_horas(row[3])
                meses = str(row[4] or '').strip()
                valor = limpar_valor(row[5])
                inser = str(row[6] or '').strip()
                link  = str(row[7] or '').strip()
                obs   = str(row[8] or '').strip()
                if add_course(admin.id, nome, 'pos', area, horas, valor, link, '', '', obs, 'ativo', inser, meses=meses):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='pos').count()} pos importados")
            break

    # ── PROFISSIONALIZANTES ───────────────────────────────────────────────
    # Cabecalho: Nº | PROFISSIONALIZANTES | ÁREA | HORAS | LINK | DESCRIÇÃO | LINK IMAGENS | OBS
    for shname in wb.sheetnames:
        if 'PROFISSIONALIZANTE' in shname.upper():
            print(f"\nImportando Profissionalizantes ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome     = row[1]
                area     = row[2]
                horas    = limpar_horas(row[3])
                link     = str(row[4] or '').strip()
                desc     = str(row[5] or '').strip()
                link_img = str(row[6] or '').strip()
                obs      = str(row[7] or '').strip()
                if add_course(admin.id, nome, 'profissionalizante', area, horas, '', link, link_img, desc, obs, 'ativo', 'INOVA'):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='profissionalizante').count()} profissionalizantes")
            break

    # ── RÁPIDOS INOVA ────────────────────────────────────────────────────
    # Cabecalho: Nº | RÁPIDOS | ÁREA | HORAS | LINKS | VALOR | DESCRIÇÃO | CAPA | OBS
    for shname in wb.sheetnames:
        if 'RÁPIDOS INOVA' in shname or 'RAPIDOS INOVA' in shname:
            print(f"\nImportando Rapidos ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome     = row[1]
                area     = row[2]
                horas    = limpar_horas(row[3])
                link     = str(row[4] or '').strip()
                valor    = limpar_valor(row[5])
                desc     = str(row[6] or '').strip()
                link_img = str(row[7] or '').strip()
                obs      = str(row[8] or '').strip()
                if add_course(admin.id, nome, 'rapido', area, horas, valor, link, link_img, desc, obs, 'ativo', 'INOVA'):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='rapido').count()} rapidos")
            break

    # ── PACOTE CURSOS ────────────────────────────────────────────────────
    # Cabecalho: Nº | PACOTE CURSOS INOVA | ÁREA | HORAS | LINKS | INSERÇOR | OBS | DATA | VALOR
    for shname in wb.sheetnames:
        if shname.upper() == 'PACOTE CURSOS':
            print(f"\nImportando Pacotes ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome  = row[1]
                area  = row[2]
                horas = limpar_horas(row[3])
                link  = str(row[4] or '').strip()
                inser = str(row[5] or '').strip()
                obs   = str(row[6] or '').strip()
                valor = limpar_valor(row[8])
                if add_course(admin.id, nome, 'pacote', area, horas, valor, link, '', '', obs, 'ativo', inser):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='pacote').count()} pacotes")
            break

    # ── CURSOS TERCEIROS ─────────────────────────────────────────────────
    # Cabecalho: Nº | CURSOS TERCEIROS | VALOR | HORAS | TIPO | ÁREA | LINK | DONO | DESCRIÇÃO | CAPA | OBS
    for shname in wb.sheetnames:
        if 'TERCEIROS' in shname.upper():
            print(f"\nImportando Terceiros ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome  = row[1]
                valor = limpar_valor(row[2])
                horas = limpar_horas(row[3])
                area  = row[5]
                link  = str(row[6] or '').strip()
                dono  = str(row[7] or '').strip()
                desc  = str(row[8] or '').strip()
                obs   = str(row[10] or '').strip()
                s = 'descontinuado' if 'DESCONTINUADO' in obs.upper() else 'ativo'
                if add_course(admin.id, nome, 'terceiros', area, horas, valor, link, '', desc, obs, s, '', dono=dono):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='terceiros').count()} terceiros")
            break

    # ── EVENTOS ──────────────────────────────────────────────────────────
    # Cabecalho: ANO | TIPO | NOME EVENTO | HORAS | LINK | DATA | VALORES | DATA CRIAÇÃO | OBS
    for shname in wb.sheetnames:
        if shname.upper() == 'EVENTOS':
            print(f"\nImportando Eventos ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[2]:
                    continue
                nome  = row[2]
                tipo_ev = str(row[1] or '').strip()  # EAD / PRESENCIAL
                horas = limpar_horas(row[3])
                link  = str(row[4] or '').strip()
                data  = str(row[5] or '').strip()
                valor = limpar_valor(row[6])
                obs   = str(row[8] or '').strip()
                # Obs inclui data e tipo do evento
                obs_completo = f"Tipo: {tipo_ev} | Data: {data}" + (f" | {obs}" if obs else "")
                if add_course(admin.id, nome, 'evento', 'EVENTO', horas, valor, link, '', '', obs_completo, 'ativo', 'INOVA'):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='evento').count()} eventos")
            break

    # ── PRÁTICAS CONECTADAS ───────────────────────────────────────────────
    # Cabecalho: ÁREA | CURSO | CUPOM | LINKS
    for shname in wb.sheetnames:
        if 'CONECTADA' in shname.upper():
            print(f"\nImportando Praticas Conectadas ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:
                    continue
                nome  = row[1]
                area  = str(row[0] or '').strip()
                cupom = str(row[2] or '').strip()
                link  = str(row[3] or '').strip()
                if add_course(admin.id, nome, 'pratica_conectada', area, '', '', link, '', '', '', 'oculto', 'INOVA', cupom=cupom):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='pratica_conectada').count()} praticas conectadas")
            break

    # ── PRÁTICAS DE ESTÁGIO ───────────────────────────────────────────────
    # Cabecalho: PRÁTICA | ÁREA | DATA | INSERSOR | CUPOM | LINK VENDA
    for shname in wb.sheetnames:
        if 'ESTAGIO' in shname.upper() or 'ESTÁGIO' in shname.upper():
            print(f"\nImportando Praticas Estagio ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                nome  = row[0]
                area  = str(row[1] or '').strip()
                inser = str(row[3] or '').strip()
                cupom = str(row[4] or '').strip()
                link  = str(row[5] or '').strip()
                if add_course(admin.id, nome, 'pratica_estagio', area, '', '', link, '', '', '', 'ativo', inser, cupom=cupom):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='pratica_estagio').count()} praticas estagio")
            break

    # ── PROJ. AMBIENTES PROF ─────────────────────────────────────────────
    # Cabecalho: Nº | PROJETOS EM AMBIENTES PROFISSIONAIS | CUPOM | LINKS | OBS
    for shname in wb.sheetnames:
        if 'AMBIENTES' in shname.upper():
            print(f"\nImportando Proj. Ambientes ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome  = row[1]
                cupom = str(row[2] or '').strip()
                link  = str(row[3] or '').strip()
                obs   = str(row[4] or '').strip()
                if add_course(admin.id, nome, 'projeto_ambiental', '', '', '', link, '', '', obs, 'oculto', 'INOVA', cupom=cupom):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='projeto_ambiental').count()} proj ambientes")
            break

    # ── GGBR - RÁPIDOS ───────────────────────────────────────────────────
    # Cabecalho: Nº | RÁPIDOS | ÁREA | HORAS | LINKS | VALOR
    for shname in wb.sheetnames:
        if 'GGBR' in shname.upper():
            print(f"\nImportando GGBR ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not is_numero(row[0]):
                    continue
                nome  = row[1]
                area  = str(row[2] or '').strip()
                horas = limpar_horas(row[3])
                link  = str(row[4] or '').strip()
                valor = limpar_valor(row[5])
                if add_course(admin.id, nome, 'ggbr', area, horas, valor, link, '', '', '', 'ativo', 'INOVA'):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='ggbr').count()} ggbr")
            break

    # ── INTEGRA EDU ──────────────────────────────────────────────────────
    # Cabecalho: NOME CURSO | HORAS | LINK | OBS
    for shname in wb.sheetnames:
        if 'INTEGRA' in shname.upper():
            print(f"\nImportando Integra Edu ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                nome  = row[0]
                horas = limpar_horas(row[1])
                link  = str(row[2] or '').strip()
                obs   = str(row[3] or '').strip()
                if add_course(admin.id, nome, 'integra_edu', '', horas, '', link, '', '', obs, 'ativo', 'INOVA'):
                    total += 1
            print(f"  -> {Course.query.filter_by(tipo='integra_edu').count()} integra edu")
            break

    db.session.commit()

    # ── CUPONS ───────────────────────────────────────────────────────────
    # Cabecalho: NOME | QUANTIDADE | DESCONTO | CURSOS | LIMITE CURSO | USO UNICO? | DATA INICIAL | DATA FINAL | OBS
    cupons = 0
    for shname in wb.sheetnames:
        if shname.upper() == 'CUPOM':
            print(f"\nImportando Cupons ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or str(row[0]).strip() in ('NOME', ''):
                    continue
                try:
                    cp = Coupon(
                        nome=str(row[0]).strip()[:100],
                        quantidade=int(row[1] or 0),
                        desconto=float(row[2] or 0),
                        cursos_tipo=str(row[3] or '').strip(),
                        limite_curso=int(row[4] or 1),
                        uso_unico=(str(row[5] or '').upper() == 'SIM'),
                        data_inicial=pd(row[6]),
                        data_final=pd(row[7]) if len(row) > 7 else None,
                        obs=str(row[8] or '').strip() if len(row) > 8 else ''
                    )
                    db.session.add(cp)
                    cupons += 1
                except Exception as e:
                    pass
            break

    # ── REEMBOLSOS ───────────────────────────────────────────────────────
    # Cabecalho: COLAB | NOME ALUNO | DATA COMPRA | SOLICITAÇÃO | VALOR | VALOR ESTORNO | NOME CURSO | CATEGORIA | ... | DATA APROVAÇÃO | MOTIVO
    reembolsos = 0
    for shname in wb.sheetnames:
        if 'REEMBOLSO' in shname.upper():
            print(f"\nImportando Reembolsos ({shname})...")
            ws = wb[shname]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1] or str(row[1]).strip() in ('NOME ALUNO', ''):
                    continue
                try:
                    r = Refund(
                        colab=str(row[0] or '').strip(),
                        nome_aluno=str(row[1] or '').strip(),
                        data_compra=pd(row[2]),
                        data_solicitacao=pd(row[3]),
                        valor=float(str(row[4] or 0).replace(',', '.') or 0),
                        valor_estorno=float(str(row[5] or 0).replace(',', '.') or 0),
                        nome_curso=str(row[6] or '').strip(),
                        categoria=str(row[7] or '').strip(),
                        data_aprovacao=pd(row[10]) if len(row) > 10 else None,
                        motivo=str(row[11] or '').strip() if len(row) > 11 else ''
                    )
                    db.session.add(r)
                    reembolsos += 1
                except Exception as e:
                    pass
            break

    db.session.commit()

    print(f"\n{'='*40}")
    print(f"IMPORTACAO CONCLUIDA!")
    print(f"{'='*40}")
    from sqlalchemy import func
    for tipo, qtd in db.session.query(Course.tipo, func.count(Course.id)).group_by(Course.tipo).order_by(Course.tipo).all():
        print(f"  {tipo:<25} {qtd} cursos")
    print(f"  {'cupons':<25} {Coupon.query.count()}")
    print(f"  {'reembolsos':<25} {Refund.query.count()}")
    print(f"  {'TOTAL CURSOS':<25} {Course.query.count()}")
