"""
Importa as matrizes curriculares dos cursos PÓS a partir da aba PÓS- MATRIZES INOVA.
NAO apaga nenhum dado existente. Só adiciona disciplinas a cursos que ainda não tenham.
"""
import sys, io, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from app import app, db, Course, Discipline
import openpyxl

def norm(s):
    """Normaliza string: maiúsculas, sem espaços duplos, expande abreviações."""
    s = re.sub(r'\s+', ' ', s.upper().strip())
    s = s.replace('ESP. EM ', 'ESPECIALIZAÇÃO EM ')
    s = s.replace('ESP.EM ', 'ESPECIALIZAÇÃO EM ')
    s = s.replace('ESPECIALIZAÇÃO EM\xa0', 'ESPECIALIZAÇÃO EM ')
    s = s.replace('ESPECIALIZAÇÃO EM', 'ESPECIALIZAÇÃO EM ')  # fix missing space
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def nome_similar(a, b, chars=25):
    """Compara os primeiros N caracteres normalizados."""
    return norm(a)[:chars] == norm(b)[:chars]

with app.app_context():
    caminho = os.path.join(os.path.dirname(__file__), 'CURSOS INOVA - LINKS (1).xlsx')
    wb = openpyxl.load_workbook(caminho)

    # ── Tabela de titulações: disciplina -> titulação ──────────────────────────
    tit_lookup = {}
    if 'TITULAÇÕES PÓS' in wb.sheetnames:
        for row in wb['TITULAÇÕES PÓS'].iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 4 and row[4]:
                key = str(row[0]).strip().upper()
                tit_lookup[key] = str(row[4]).strip()

    # ── Ler cursos PÓS do banco ────────────────────────────────────────────────
    pos_courses = Course.query.filter_by(tipo='pos').all()
    print(f'Cursos POS no banco: {len(pos_courses)}')

    # ── Processar aba PÓS- MATRIZES INOVA ─────────────────────────────────────
    sheet_name = next((s for s in wb.sheetnames if 'MATRIZES' in s.upper()), None)
    if not sheet_name:
        print('[ERRO] Aba POS-MATRIZES nao encontrada.')
        sys.exit(1)

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(min_row=3, values_only=True))

    # Encontrar onde termina a lista de cursos (último row numérico)
    matrix_start = len(all_rows)
    for i, row in enumerate(all_rows):
        try:
            int(str(row[0] or '').strip())
        except:
            matrix_start = i
            break

    print(f'Secao de matrizes começa no offset {matrix_start} (linha {matrix_start + 3})\n')

    current_course_id = None
    current_course_name = ''
    in_disciplines = False
    disc_ordem = 0
    current_modulo = None
    total_disc = 0
    skipped_courses = []
    matched_courses = []

    for row in all_rows[matrix_start:]:
        col0 = str(row[0] or '').strip() if row[0] is not None else ''
        col1 = str(row[1] or '').strip() if row[1] is not None else ''
        col2 = str(row[2] or '').strip() if row[2] is not None else ''
        col3 = str(row[3] or '').strip() if row[3] is not None else ''

        # Linha totalmente vazia → reset
        if not col0 and not col1 and not col2 and not col3:
            in_disciplines = False
            continue

        # Cabeçalho de disciplinas (DISCIPLINAS, CH, PROFESSORES)
        if col1.upper() == 'DISCIPLINAS':
            in_disciplines = True
            disc_ordem = 0
            current_modulo = None
            continue

        # Linha de total (sem col0, sem col1, tem col2 = total de horas)
        if not col0 and not col1 and col2:
            continue

        # Cabeçalho de curso:
        #   Tipo A: col0=nome, col1=vazio, col2=vazio
        #   Tipo B: col0=vazio, col1=nome, col2=vazio
        if not col2:
            course_name = col0 if (col0 and not col1) else col1
            if course_name and course_name.upper() not in ('PROFESSOR', 'PROFESSORES', 'DISCIPLINAS'):
                in_disciplines = False
                disc_ordem = 0
                current_modulo = None
                current_course_id = None
                current_course_name = course_name

                # Tentar casar com curso no banco
                match = None
                for c in pos_courses:
                    if nome_similar(c.nome, course_name, 25):
                        match = c
                        break
                # Se não achou com 25 chars, tentar com 15
                if not match:
                    for c in pos_courses:
                        if nome_similar(c.nome, course_name, 15):
                            match = c
                            break

                if match:
                    # Pular se já tem disciplinas cadastradas
                    existing = Discipline.query.filter_by(course_id=match.id).count()
                    if existing > 0:
                        skipped_courses.append(f'{match.nome[:50]} (ja tem {existing} disc.)')
                        current_course_id = None
                    else:
                        current_course_id = match.id
                        matched_courses.append(match.nome[:50])
                else:
                    print(f'  [!] Nao encontrou curso para: {course_name[:60]}')
                    current_course_id = None
            continue

        # Linha de disciplina: tem col1 (nome) e col2 (CH)
        if in_disciplines and col1 and col2 and current_course_id:
            # Atualizar módulo/ordem
            if col0.startswith('Mód') or (col0 and not col0.isdigit() and not col0.startswith('Mód') == False):
                # col0 é um módulo como "Mód. 01"
                if col0 and not col0.isdigit():
                    current_modulo = col0
                    disc_ordem += 1
                else:
                    try:
                        disc_ordem = int(col0)
                    except:
                        disc_ordem += 1
            elif col0 and col0.isdigit():
                disc_ordem = int(col0)
            else:
                disc_ordem += 1

            tit = tit_lookup.get(col1.upper().strip(), '')
            d = Discipline(
                course_id=current_course_id,
                modulo=current_modulo or None,
                ordem=disc_ordem,
                nome=col1,
                carga=col2 or None,
                professor=col3 or None,
                titulacao=tit or None,
            )
            db.session.add(d)
            total_disc += 1

    db.session.commit()

    print(f'\n{"="*50}')
    print(f'MATRIZES IMPORTADAS: {total_disc} disciplinas')
    print(f'Cursos com matriz adicionada: {len(matched_courses)}')
    for n in matched_courses:
        print(f'  + {n}')
    if skipped_courses:
        print(f'\nCursos ja tinham disciplinas (ignorados): {len(skipped_courses)}')
        for n in skipped_courses:
            print(f'  ~ {n}')
    print(f'{"="*50}')
    print(f'Total disciplinas no banco: {Discipline.query.count()}')
