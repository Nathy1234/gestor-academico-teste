import openpyxl
wb = openpyxl.load_workbook('CURSOS INOVA - LINKS (1).xlsx')

# Profissionalizantes - linhas apos os cursos (após linha 44)
ws = wb['PROFISSIONALIZANTES INOVA']
print(f'=== PROFISSIONALIZANTES - linhas 44 a 120 (total={ws.max_row}) ===')
for i, row in enumerate(ws.iter_rows(min_row=44, max_row=120, values_only=True), 44):
    vals = [str(c or '')[:60] for c in row[:9]]
    if any(v.strip() for v in vals):
        print(f'L{i}: {vals}')

print()
# POS - estrutura das matrizes (após os cursos)
ws2 = wb['PÓS- MATRIZES INOVA']
# Contar quantos cursos há (linhas com número na col 0)
cursos_pos = []
for i, row in enumerate(ws2.iter_rows(min_row=3, values_only=True), 3):
    try:
        num = int(str(row[0] or '').strip())
        cursos_pos.append((i, num, str(row[1] or '')[:50]))
    except:
        pass

print(f'=== POS: {len(cursos_pos)} cursos encontrados ===')
for linha, num, nome in cursos_pos[:5]:
    print(f'  L{linha} #{num}: {nome}')
print(f'  ... ultimo: L{cursos_pos[-1][0]} #{cursos_pos[-1][1]}: {cursos_pos[-1][2]}')

print()
# Linhas apos os cursos de pos (estrutura de matrizes)
ultima_linha_curso = cursos_pos[-1][0] if cursos_pos else 50
print(f'=== POS MATRIZES: primeiras 30 linhas apos cursos (a partir L{ultima_linha_curso+1}) ===')
count = 0
for i, row in enumerate(ws2.iter_rows(min_row=ultima_linha_curso+1, values_only=True), ultima_linha_curso+1):
    vals = [str(c or '')[:60] for c in row[:6]]
    if any(v.strip() for v in vals):
        print(f'L{i}: {vals}')
        count += 1
        if count >= 30:
            break

print()
# TITULACOES - ver estrutura completa
for name in wb.sheetnames:
    if 'TITULA' in name.upper():
        ws3 = wb[name]
        print(f'=== {name}: primeiras 5 e ultimas 5 linhas ===')
        rows = list(ws3.iter_rows(min_row=1, max_row=5, values_only=True))
        for i, row in enumerate(rows, 1):
            print(f'L{i}: {[str(c or "")[:50] for c in row[:5]]}')
        print('...')
        all_rows = list(ws3.iter_rows(values_only=True))
        for i, row in enumerate(all_rows[-3:], len(all_rows)-2):
            print(f'L{i}: {[str(c or "")[:50] for c in row[:5]]}')
        break
