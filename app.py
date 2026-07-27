from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from datetime import datetime, timedelta
from functools import wraps
import hashlib, os, secrets, shutil, json, threading, time, io, zipfile, unicodedata as _ucd

def _norm_name(s):
    """Remove acentos e converte para maiúsculo — para comparação de nomes de insersores."""
    return ''.join(c for c in _ucd.normalize('NFD', s.upper()) if _ucd.category(c) != 'Mn')

# Mapa de iniciais usadas no campo insersor (ex: 'P,L,S,F') para nome canônico normalizado
INICIAIS_INSERCAO = {
    'N': 'NATALIA', 'P': 'PEDRO', 'L': 'LUCAS',
    'S': 'STEFANYE', 'F': 'FELIPE', 'J': 'JUNIOR',
}

# Carrega variáveis do .env manualmente para não depender do python-dotenv
# e sem ativar DATABASE_URL (mantém SQLite local)
def _load_env_var(key):
    try:
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
        with open(env_path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('#') or '=' not in line:
                    continue
                k, _, v = line.partition('=')
                if k.strip() == key:
                    return v.strip().strip('"').strip("'")
    except Exception:
        pass
    return None

# Carrega do .env (se não estiverem no ambiente) apenas as chaves de serviços
# externos — nunca SECRET_KEY/DATABASE_URL, pra manter SQLite local por padrão.
for _chave in ('ANTHROPIC_API_KEY', 'EMAIL_SMTP_USER', 'EMAIL_SMTP_PASSWORD'):
    if not os.environ.get(_chave):
        _val = _load_env_var(_chave)
        if _val and _val != 'sua-chave-aqui':
            os.environ[_chave] = _val

app = Flask(__name__)

AREAS_VALIDAS = [
    'EDUCAÇÃO', 'SAÚDE', 'NEGÓCIOS', 'TECNOLOGIA',
    'CRIATIVIDADE', 'GASTRONOMIA', 'EVENTO',
]
TIPOS_CURSO = [
    'pos', 'profissionalizante', 'rapido', 'pacote', 'terceiros',
    'evento', 'pratica_conectada', 'pratica_estagio',
    'projeto_ambiental', 'ggbr', 'integra_edu',
]

def responsaveis_atuais():
    """Responsáveis por inserção de cursos = usuários cadastrados no sistema
    (substituiu a lista fixa antiga — para adicionar/remover alguém, basta
    criar ou excluir a conta em Usuários, sem precisar mexer no código)."""
    return [u.username for u in User.query.order_by(User.username).all()]

def _insersor_contains(insersor_field, username):
    """Verifica se `username` está entre os insersores de um curso, aceitando
    tanto o nome completo quanto as iniciais legadas (ex: 'N' = Natália)."""
    if not insersor_field:
        return False
    un_norm = _norm_name(username)
    inicial = next((k for k, v in INICIAIS_INSERCAO.items() if v == un_norm), None)
    for parte in insersor_field.split(','):
        p = parte.strip()
        if _norm_name(p) == un_norm:
            return True
        if inicial and len(p) == 1 and p.upper() == inicial:
            return True
    return False

EMAIL_DOMINIO_PERMITIDO = '@unifatecie.edu.br'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
if not os.environ.get('SECRET_KEY'):
    print('[AVISO] SECRET_KEY não definida nas variáveis de ambiente. '
          'Gerando uma chave temporária para esta execução — os usuários serão '
          'desconectados a cada reinício do servidor. Defina SECRET_KEY no ambiente '
          '(Vercel/host) para sessões estáveis e seguras.')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///inova.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
# Em produção (Postgres) o cookie de sessão só trafega em HTTPS; em SQLite
# local (desenvolvimento) isso quebraria o login via http://localhost.
app.config['SESSION_COOKIE_SECURE'] = _db_url.startswith('postgresql://')
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
if _db_url.startswith('postgresql://'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'connect_args': {
            'sslmode': 'require',
            'connect_timeout': 10,
        },
        'pool_pre_ping': True,
        'pool_size': 1,
        'max_overflow': 0,
        'pool_timeout': 20,
        'pool_recycle': 300,
    }
db = SQLAlchemy(app)
csrf = CSRFProtect(app)
limiter = Limiter(get_remote_address, app=app, storage_uri='memory://', default_limits=[])

# ─── E-MAIL ────────────────────────────────────────────────────────────────────
EMAIL_SMTP_USER = os.environ.get('EMAIL_SMTP_USER')
EMAIL_SMTP_PASSWORD = os.environ.get('EMAIL_SMTP_PASSWORD')

def enviar_email(destinatario, assunto, texto):
    """Envia e-mail via Gmail SMTP. Se EMAIL_SMTP_USER/PASSWORD não estiverem
    configurados, não falha — só registra no console, o que permite testar os
    fluxos de e-mail localmente sem precisar da conta configurada ainda."""
    if not destinatario:
        return False
    if not (EMAIL_SMTP_USER and EMAIL_SMTP_PASSWORD):
        print(f'[EMAIL SIMULADO — EMAIL_SMTP_USER/PASSWORD não configurados]\n'
              f'Para: {destinatario}\nAssunto: {assunto}\n\n{texto}\n')
        return True
    try:
        import smtplib
        from email.mime.text import MIMEText
        msg = MIMEText(texto, 'plain', 'utf-8')
        msg['Subject'] = assunto
        msg['From'] = f'Gestor Acadêmico <{EMAIL_SMTP_USER}>'
        msg['To'] = destinatario
        with smtplib.SMTP('smtp.gmail.com', 587, timeout=10) as server:
            server.starttls()
            server.login(EMAIL_SMTP_USER, EMAIL_SMTP_PASSWORD)
            server.sendmail(EMAIL_SMTP_USER, [destinatario], msg.as_string())
        return True
    except Exception as e:
        print(f'[ERRO EMAIL] {e}')
        return False

def _reset_senha_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='reset-senha')

# ─── MODELS ────────────────────────────────────────────────────────────────────

class User(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    username     = db.Column(db.String(80), unique=True, nullable=False)
    nome         = db.Column(db.String(200))  # nome completo, para exibição em e-mails
    email        = db.Column(db.String(200))
    password     = db.Column(db.String(200), nullable=False)
    must_change_password = db.Column(db.Boolean, default=False)
    role         = db.Column(db.String(20), default='viewer')  # admin, editor, viewer
    permissoes   = db.Column(db.Text, default='{}')  # JSON com permissoes especificas
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)

    def get_perm(self, key):
        if self.role == 'admin':
            return True
        try:
            p = json.loads(self.permissoes or '{}')
        except:
            p = {}
        return p.get(key, False)

    def _p(self):
        try:
            return json.loads(self.permissoes or '{}')
        except:
            return {}

    def can_edit(self):
        return self.role in ('admin', 'editor') or self.get_perm('cursos_editar')

    def can_delete(self):
        return self.role == 'admin' or self.get_perm('cursos_excluir')

    def can_manage_cupons(self):
        if self._p().get('block_cupons'): return False
        return True  # todos os usuários logados têm acesso por padrão

    def can_manage_reembolsos(self):
        if self._p().get('block_reembolsos'): return False
        return True  # todos os usuários logados têm acesso por padrão

    def can_view_historico(self):
        if self._p().get('block_historico'): return False
        return True  # todos os usuários logados têm acesso por padrão

    def can_manage_usuarios(self):
        return self.role == 'admin' or self.get_perm('usuarios_gerenciar')

    def can_manage_backup(self):
        return self.role == 'admin' or self.get_perm('backup_gerenciar')

    def can_change_own_password(self):
        if self.role == 'admin': return True
        return not self._p().get('block_trocar_senha')

class Course(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    nome          = db.Column(db.String(300), nullable=False)
    tipo          = db.Column(db.String(50))   # pos, profissionalizante, rapido, pacote, terceiros, evento, pratica, projeto
    area          = db.Column(db.String(100))
    horas         = db.Column(db.String(20))
    meses         = db.Column(db.String(20))
    valor         = db.Column(db.String(50))
    link_venda    = db.Column(db.Text)
    descricao     = db.Column(db.Text)
    link_imagem   = db.Column(db.Text)
    insersor      = db.Column(db.Text)  # pode ser comma-separated para múltiplos insersores
    obs           = db.Column(db.Text)
    status        = db.Column(db.String(30), default='ativo')  # ativo, descontinuado, em_edicao, oculto
    cupom         = db.Column(db.String(100))
    dono          = db.Column(db.Text)        # para cursos terceiros
    ano           = db.Column(db.String(10))  # ano de criação/edição do curso
    extra_data    = db.Column(db.Text)        # JSON com dados extras (matriz, disciplinas, etc.)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    created_by    = db.Column(db.Integer, db.ForeignKey('user.id'))

class Discipline(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    modulo    = db.Column(db.String(100))
    ordem     = db.Column(db.Integer)
    nome      = db.Column(db.String(300), nullable=False)
    carga     = db.Column(db.String(20))
    professor = db.Column(db.String(200))
    cod_moodle    = db.Column(db.String(50))
    titulacao     = db.Column(db.String(50))
    plataforma_ok = db.Column(db.Boolean, default=False)
    plataforma_em = db.Column(db.DateTime)

class AuditLog(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'))
    username   = db.Column(db.String(80))
    action     = db.Column(db.String(50))   # criar, editar, excluir, login
    entity     = db.Column(db.String(50))   # course, user, cupom
    entity_id  = db.Column(db.Integer)
    detail     = db.Column(db.Text)         # JSON do que mudou
    timestamp  = db.Column(db.DateTime, default=datetime.utcnow)

class Coupon(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    nome         = db.Column(db.String(100), nullable=False)
    quantidade   = db.Column(db.Integer)
    desconto     = db.Column(db.Float)
    cursos_tipo  = db.Column(db.String(100))
    limite_curso = db.Column(db.Integer)
    uso_unico    = db.Column(db.Boolean, default=True)
    data_inicial = db.Column(db.Date)
    data_final   = db.Column(db.Date)
    obs          = db.Column(db.Text)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)

class Refund(db.Model):
    id               = db.Column(db.Integer, primary_key=True)
    colab            = db.Column(db.String(100))
    nome_aluno       = db.Column(db.String(200))
    data_compra      = db.Column(db.Date)
    data_solicitacao = db.Column(db.Date)   # Solicitação do aluno
    valor            = db.Column(db.Float)
    valor_estorno    = db.Column(db.Float)
    nome_curso       = db.Column(db.String(300))
    categoria        = db.Column(db.String(100))
    solicitacao_1    = db.Column(db.Date)   # 1ª Solicitação (Nathy)
    solicitacao_2    = db.Column(db.Date)   # 2ª Solicitação (Jorge)
    data_aprovacao   = db.Column(db.Date)
    motivo           = db.Column(db.Text)
    curso_excluido   = db.Column(db.Date)   # Data exclusão do curso
    obs              = db.Column(db.Text)
    concluido_manual = db.Column(db.Boolean, default=False)
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    # Dados para pagamento do reembolso
    cpf              = db.Column(db.String(20))
    celular          = db.Column(db.String(30))
    pix              = db.Column(db.String(200))
    email_destino    = db.Column(db.String(200))

    @property
    def pendencia(self):
        if self.concluido_manual:
            return ('concluido', 'Concluído')
        if not self.solicitacao_1:
            return ('sem_solic1', 'Aguarda 1ª Solicitação')
        if not self.solicitacao_2:
            return ('sem_solic2', 'Aguarda 2ª Solicitação')
        if not self.data_aprovacao:
            return ('sem_aprovacao', 'Aguarda Aprovação')
        if not self.curso_excluido:
            return ('sem_exclusao', 'Aguarda Exclusão do Curso')
        return ('concluido', 'Concluído')

class BackupRecord(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    filename   = db.Column(db.String(200))
    size_kb    = db.Column(db.Float)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    tipo       = db.Column(db.String(20), default='auto')  # auto, manual

# ─── HELPERS ───────────────────────────────────────────────────────────────────

def hash_pw(pw): return generate_password_hash(pw)

def check_pw(stored_hash, plain_pw):
    """Verifica a senha. Aceita hashes novos (werkzeug, com salt) e os
    hashes antigos em SHA-256 puro criados antes desta correção de segurança."""
    if stored_hash.startswith(('pbkdf2:', 'scrypt:')):
        return check_password_hash(stored_hash, plain_pw)
    return stored_hash == hashlib.sha256(plain_pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        u = User.query.get(session['user_id'])
        if not u or u.role != 'admin':
            flash('Acesso restrito a administradores.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def editor_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        u = User.query.get(session['user_id'])
        if not u or u.role not in ('admin', 'editor'):
            flash('Sem permissão para editar.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def perm_check(check_fn_name):
    """Decorator que verifica uma permissão específica via método do User model."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            u = User.query.get(session['user_id'])
            if not u or not getattr(u, check_fn_name)():
                flash('Você não tem permissão para acessar esta seção.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def log_action(user_id, username, action, entity, entity_id, detail=''):
    entry = AuditLog(user_id=user_id, username=username, action=action,
                     entity=entity, entity_id=entity_id, detail=detail)
    db.session.add(entry)
    db.session.commit()

def make_backup():
    with app.app_context():
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        src = os.path.join(app.instance_path, 'inova.db')
        if not os.path.exists(src): return
        fname = f'backup_{ts}.zip'
        fpath = os.path.join('backups', fname)
        os.makedirs('backups', exist_ok=True)
        with zipfile.ZipFile(fpath, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(src, 'inova.db')
        size_kb = os.path.getsize(fpath) / 1024
        rec = BackupRecord(filename=fname, size_kb=round(size_kb, 2), tipo='auto')
        db.session.add(rec)
        db.session.commit()
        # Keep only last 20 backups
        bks = BackupRecord.query.order_by(BackupRecord.created_at.asc()).all()
        if len(bks) > 20:
            for old in bks[:-20]:
                try: os.remove(os.path.join('backups', old.filename))
                except: pass
                db.session.delete(old)
            db.session.commit()

def backup_scheduler():
    while True:
        time.sleep(7 * 24 * 3600)  # weekly
        make_backup()

@app.context_processor
def inject_now():
    return {'now': datetime.now}

@app.context_processor
def inject_notificacoes():
    if 'user_id' not in session:
        return {}
    u = User.query.get(session['user_id'])
    if not u:
        return {}
    # Disciplinas pendentes (plataforma_ok=False) em cursos atribuídos a este usuário
    q = db.session.query(Discipline, Course)\
        .join(Course, Discipline.course_id == Course.id)\
        .filter(Discipline.plataforma_ok == False)\
        .filter(Course.status.notin_(['descontinuado']))\
        .filter(Course.insersor != None, Course.insersor != '')
    rows = q.order_by(Course.nome, Discipline.ordem).all()
    if u.role != 'admin':
        # Filtra em Python (não em SQL) para reconhecer também as iniciais
        # legadas (ex: curso com insersor='N' deve notificar a Natália).
        rows = [(disc, curso) for disc, curso in rows if _insersor_contains(curso.insersor, u.username)]
    pendentes = rows

    by_course = {}
    for disc, curso in pendentes:
        if curso.id not in by_course:
            by_course[curso.id] = {'course': curso, 'discs': [], 'total': 0}
        by_course[curso.id]['discs'].append(disc)
        by_course[curso.id]['total'] += 1

    notif_list = sorted(by_course.values(), key=lambda x: x['total'], reverse=True)[:8]

    # Para admin: cursos marcados como finalizado aguardando publicação
    admin_finalizado = []
    if u.role == 'admin':
        admin_finalizado = Course.query.filter_by(status='finalizado')\
            .order_by(Course.updated_at.desc()).limit(15).all()

    return {
        'notif_count': len(pendentes),
        'notif_list': notif_list,
        'admin_finalizado': admin_finalizado,
        'admin_finalizado_count': len(admin_finalizado),
        'can_cupons': u.can_manage_cupons(),
        'can_reembolsos': u.can_manage_reembolsos(),
        'can_historico': u.can_view_historico(),
    }

# ─── AUTH ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' in session: return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
@limiter.limit("10 per minute", methods=['POST'])
def login():
    if request.method == 'POST':
        identificador = request.form.get('email', '').strip().lower()
        # Aceita e-mail institucional (novo padrão) ou nome de usuário (contas
        # antigas ainda sem e-mail cadastrado) até que todas as contas migrem.
        u = User.query.filter(
            db.or_(db.func.lower(User.email) == identificador,
                   db.func.lower(User.username) == identificador)
        ).first()
        if u and check_pw(u.password, request.form['password']):
            if not u.password.startswith(('pbkdf2:', 'scrypt:')):
                u.password = hash_pw(request.form['password'])
                db.session.commit()
            session.permanent = True
            session['user_id'] = u.id
            session['username'] = u.username
            session['role'] = u.role
            log_action(u.id, u.username, 'login', 'user', u.id)
            return redirect(url_for('dashboard'))
        flash('Usuário ou senha incorretos.', 'danger')
    return render_template('login.html')

@app.route('/esqueci-senha', methods=['GET','POST'])
@limiter.limit("5 per minute", methods=['POST'])
def esqueci_senha():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        u = User.query.filter(db.func.lower(User.email) == email).first()
        if u:
            token = _reset_senha_serializer().dumps(u.id)
            link = url_for('resetar_senha', token=token, _external=True)
            corpo = (
                f'Olá {u.nome or u.username},\n\n'
                f'Recebemos um pedido para redefinir sua senha no Gestor Acadêmico.\n'
                f'Clique no link abaixo para escolher uma nova senha (válido por 1 hora):\n\n'
                f'{link}\n\n'
                f'Se você não pediu essa redefinição, pode ignorar este e-mail.'
            )
            enviar_email(u.email, 'Redefinição de senha — Gestor Acadêmico', corpo)
        # Mensagem sempre igual, exista ou não o e-mail — evita confirmar pra quem
        # está tentando descobrir quais e-mails têm conta no sistema.
        flash('Se esse e-mail estiver cadastrado, enviamos um link de redefinição.', 'success')
        return redirect(url_for('login'))
    return render_template('esqueci_senha.html')

@app.route('/resetar-senha/<token>', methods=['GET','POST'])
def resetar_senha(token):
    try:
        user_id = _reset_senha_serializer().loads(token, max_age=3600)
    except (BadSignature, SignatureExpired):
        flash('Link inválido ou expirado. Solicite a redefinição novamente.', 'danger')
        return redirect(url_for('esqueci_senha'))
    u = User.query.get(user_id)
    if not u:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('esqueci_senha'))
    if request.method == 'POST':
        nova = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')
        if len(nova) < 8:
            flash('A nova senha precisa ter pelo menos 8 caracteres.', 'danger')
        elif nova != confirmar:
            flash('As senhas não coincidem.', 'danger')
        else:
            u.password = hash_pw(nova)
            u.must_change_password = False
            db.session.commit()
            log_action(u.id, u.username, 'resetar_senha_email', 'user', u.id)
            flash('Senha redefinida com sucesso! Faça login com a nova senha.', 'success')
            return redirect(url_for('login'))
    return render_template('resetar_senha.html', token=token)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/minha-conta', methods=['GET','POST'])
@login_required
def minha_conta():
    u = User.query.get(session['user_id'])
    if request.method == 'POST':
        if not u.can_change_own_password():
            flash('Você não tem permissão para alterar sua própria senha. Fale com um administrador.', 'danger')
            return redirect(url_for('minha_conta'))
        senha_atual = request.form.get('senha_atual', '')
        nova = request.form.get('nova_senha', '')
        confirmar = request.form.get('confirmar_senha', '')
        if not check_pw(u.password, senha_atual):
            flash('Senha atual incorreta.', 'danger')
        elif len(nova) < 8:
            flash('A nova senha precisa ter pelo menos 8 caracteres.', 'danger')
        elif nova != confirmar:
            flash('As senhas novas não coincidem.', 'danger')
        else:
            u.password = hash_pw(nova)
            u.must_change_password = False
            db.session.commit()
            log_action(u.id, u.username, 'trocar_senha', 'user', u.id)
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('dashboard'))
    return render_template('minha_conta.html', u=u)

# ─── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    from sqlalchemy import or_ as sql_or, func as sql_func

    u = User.query.get(session['user_id'])
    is_admin = u.role == 'admin'
    filtro_ins = request.args.get('insersor', '')

    # Para não-admins, aplica filtro automático pelo nome do próprio usuário
    def _ins_filter(q, nome):
        n = nome.lower()
        n_norm = _norm_name(nome)
        inicial = next((k for k, v in INICIAIS_INSERCAO.items() if v == n_norm), None)
        conds = [
            sql_func.lower(Course.insersor) == n,
            sql_func.lower(Course.insersor).like(f'{n},%'),
            sql_func.lower(Course.insersor).like(f'%,{n}'),
            sql_func.lower(Course.insersor).like(f'%,{n},%'),
        ]
        if inicial:
            i = inicial.lower()
            conds += [
                sql_func.lower(Course.insersor) == i,
                sql_func.lower(Course.insersor).like(f'{i},%'),
                sql_func.lower(Course.insersor).like(f'%,{i}'),
                sql_func.lower(Course.insersor).like(f'%,{i},%'),
            ]
        return q.filter(sql_or(*conds))

    q_base = Course.query
    if is_admin and filtro_ins:
        q_base = _ins_filter(q_base, filtro_ins)
    elif not is_admin:
        q_base = _ins_filter(q_base, u.username)

    total      = q_base.count()
    ativos     = q_base.filter_by(status='ativo').count()
    em_edicao  = q_base.filter_by(status='em_edicao').count()
    desc       = q_base.filter_by(status='descontinuado').count()
    ocultos    = q_base.filter_by(status='oculto').count()
    finalizado = q_base.filter_by(status='finalizado').count()

    por_tipo = db.session.query(Course.tipo, db.func.count(Course.id))\
                         .group_by(Course.tipo).all()

    if is_admin:
        recentes = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(10).all()
    else:
        recentes = AuditLog.query.filter(
            sql_func.lower(AuditLog.username) == u.username.lower()
        ).order_by(AuditLog.timestamp.desc()).limit(10).all()

    ultimo_bk = BackupRecord.query.order_by(BackupRecord.created_at.desc()).first()

    # Andamento por insersor: pendentes e concluídas por pessoa
    # Para não-admins: apenas a própria linha
    equipe = [n.upper() for n in responsaveis_atuais()] if is_admin else [u.username.upper()]

    stats_map = {nome.upper(): {'pendentes': 0, 'concluidas': 0, 'total': 0} for nome in equipe}

    raw_disc = db.session.query(Course.insersor, Discipline.plataforma_ok, db.func.count(Discipline.id))\
        .join(Discipline, Discipline.course_id == Course.id)\
        .filter(Course.insersor != None, Course.insersor != '')\
        .group_by(Course.insersor, Discipline.plataforma_ok).all()

    for ins_field, ok, qtd in raw_disc:
        for parte in ins_field.split(','):
            p = parte.strip()
            p_norm = _norm_name(p)
            # Expande inicial única (ex: 'S' → 'STEFANYE')
            if len(p) == 1:
                p_norm = INICIAIS_INSERCAO.get(p.upper(), p_norm)
            for canonical in equipe:
                if p_norm == _norm_name(canonical):
                    stats_map[canonical.upper()]['total'] += qtd
                    if ok:
                        stats_map[canonical.upper()]['concluidas'] += qtd
                    else:
                        stats_map[canonical.upper()]['pendentes'] += qtd
                    break

    pend_por_ins = sorted(
        [(nome, v['pendentes'], v['concluidas'], v['total'])
         for nome, v in stats_map.items() if v['total'] > 0 or is_admin],
        key=lambda x: x[1], reverse=True
    )

    insersores = responsaveis_atuais() if is_admin else []

    # Card: cursos por responsável (insersor)
    cursos_ins_stats = []
    nomes_ins = responsaveis_atuais() if is_admin else [u.username]
    for ins_nome in nomes_ins:
        q_ins = _ins_filter(Course.query, ins_nome)
        total_ins = q_ins.count()
        if not is_admin and total_ins == 0:
            continue
        ativos_ins   = q_ins.filter_by(status='ativo').count()
        em_ed_ins    = q_ins.filter_by(status='em_edicao').count()
        ids_ins = [r[0] for r in q_ins.with_entities(Course.id).all()]
        pend_disc_ins = Discipline.query.filter(
            Discipline.course_id.in_(ids_ins),
            Discipline.plataforma_ok == False
        ).count() if ids_ins else 0
        cursos_ins_stats.append({
            'nome': ins_nome,
            'total': total_ins,
            'ativos': ativos_ins,
            'em_edicao': em_ed_ins,
            'pend_disc': pend_disc_ins,
        })

    return render_template('dashboard.html',
        total=total, ativos=ativos, em_edicao=em_edicao, desc=desc,
        ocultos=ocultos, finalizado=finalizado,
        por_tipo=por_tipo, recentes=recentes,
        ultimo_bk=ultimo_bk, pend_por_ins=pend_por_ins,
        insersores=insersores, filtro_ins=filtro_ins,
        is_admin=is_admin, usuario_atual=u,
        cursos_ins_stats=cursos_ins_stats)

# ─── COURSES ───────────────────────────────────────────────────────────────────

@app.route('/cursos')
@login_required
def cursos():
    tipo     = request.args.get('tipo','')
    area     = request.args.get('area','')
    status   = request.args.get('status','')
    busca    = request.args.get('q','')
    insersor = request.args.get('insersor','')
    lista = _build_cursos_query(tipo, area, status, busca, insersor)
    areas  = AREAS_VALIDAS
    insersores = [u.username for u in User.query.order_by(User.username).all()]
    return render_template('cursos.html', cursos=lista, areas=areas, insersores=insersores,
                           filtro_tipo=tipo, filtro_area=area, filtro_status=status,
                           filtro_insersor=insersor, busca=busca)

def _build_cursos_query(tipo, area, status, busca, insersor):
    from sqlalchemy import or_ as sql_or, func as sql_func
    q = Course.query
    if tipo:     q = q.filter_by(tipo=tipo)
    if area:     q = q.filter_by(area=area)
    if status:   q = q.filter_by(status=status)
    if insersor:
        ins = insersor.lower()
        q = q.filter(sql_or(
            sql_func.lower(Course.insersor) == ins,
            sql_func.lower(Course.insersor).like(f'{ins},%'),
            sql_func.lower(Course.insersor).like(f'%,{ins}'),
            sql_func.lower(Course.insersor).like(f'%,{ins},%'),
        ))
    if busca:    q = q.filter(Course.nome.ilike(f'%{busca}%'))
    return q.order_by(Course.nome).all()

@app.route('/cursos/exportar-excel')
@login_required
def cursos_exportar_excel():
    import openpyxl, re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    tipo     = request.args.get('tipo','')
    area     = request.args.get('area','')
    status   = request.args.get('status','')
    busca    = request.args.get('q','')
    insersor = request.args.get('insersor','')
    lista = _build_cursos_query(tipo, area, status, busca, insersor)

    TIPO_LABELS = {
        'pos':'Pós-Graduação','profissionalizante':'Profissionalizante','rapido':'Rápido',
        'pacote':'Pacote','terceiros':'Terceiros','evento':'Evento',
        'pratica_conectada':'Prática Conectada','pratica_estagio':'Prática Estágio',
        'projeto_ambiental':'Proj. Ambiental','ggbr':'GGBR','integra_edu':'Integra Edu',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cursos'

    thin  = Side(style='thin', color='CBD5E1')
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap  = Alignment(wrap_text=True, vertical='top')
    center = Alignment(horizontal='center', vertical='center')
    hfill = PatternFill('solid', fgColor='6366F1')
    hfont = Font(bold=True, color='FFFFFF', size=10)

    STATUS_COLORS = {
        'ativo':'DCFCE7','em_edicao':'FEF9C3','descontinuado':'F1F5F9',
        'oculto':'EDE9FE','finalizado':'DBEAFE','externo':'FFEDD5',
    }

    headers = ['Nome do Curso','Tipo','Área','Status','CH','Duração','Valor (R$)',
               'Insersor','Cupom','Ano','Dono/Professor','Link Venda','Obs']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr
    ws.row_dimensions[1].height = 28

    for i, curso in enumerate(lista, 2):
        sc = STATUS_COLORS.get(curso.status, 'FFFFFF')
        row_fill = PatternFill('solid', fgColor=sc)
        vals = [
            curso.nome,
            TIPO_LABELS.get(curso.tipo, curso.tipo),
            curso.area or '',
            curso.status.replace('_',' ').title(),
            curso.horas or '',
            curso.meses or '',
            curso.valor or '',
            curso.insersor or '',
            curso.cupom or '',
            curso.ano or '',
            curso.dono or '',
            curso.link_venda or '',
            curso.obs or '',
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = bdr; cell.alignment = wrap; cell.fill = row_fill

    widths = [55,18,12,14,8,10,10,20,14,6,18,45,30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from datetime import datetime as dt
    fname = f'cursos_inova_{dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/cursos/relatorio')
@login_required
def cursos_relatorio():
    tipo     = request.args.get('tipo','')
    area     = request.args.get('area','')
    status   = request.args.get('status','')
    busca    = request.args.get('q','')
    insersor = request.args.get('insersor','')
    lista = _build_cursos_query(tipo, area, status, busca, insersor)
    insersores = [u.username for u in User.query.order_by(User.username).all()]
    return render_template('cursos_relatorio.html', cursos=lista,
                           filtro_tipo=tipo, filtro_area=area, filtro_status=status,
                           filtro_insersor=insersor, busca=busca,
                           areas=AREAS_VALIDAS, insersores=insersores,
                           now=datetime.utcnow())

@app.route('/cursos/novo', methods=['GET','POST'])
@editor_required
def curso_novo():
    if request.method == 'POST':
        d = request.form
        extra = {}
        # matrix fields passed as JSON string
        if d.get('disciplinas_json'):
            extra['disciplinas'] = json.loads(d['disciplinas_json'])
        c = Course(
            nome=d['nome'], tipo=d['tipo'], area=d.get('area',''),
            horas=d.get('horas',''), meses=d.get('meses',''), valor=d.get('valor',''),
            link_venda=d.get('link_venda',''), descricao=d.get('descricao',''),
            link_imagem=d.get('link_imagem',''), insersor=','.join(d.getlist('insersores')) or session['username'],
            obs=d.get('obs',''), status=d.get('status','em_edicao'),
            ano=d.get('ano',''), cupom=d.get('cupom',''), dono=d.get('dono',''),
            extra_data=json.dumps(extra, ensure_ascii=False),
            created_by=session['user_id']
        )
        db.session.add(c)
        db.session.flush()
        # Save disciplines separately
        discs = json.loads(d.get('disciplinas_json','[]') or '[]')
        for i, disc in enumerate(discs):
            dd = Discipline(course_id=c.id, modulo=disc.get('modulo',''),
                           ordem=i+1, nome=disc.get('nome',''),
                           carga=disc.get('carga',''), professor=disc.get('professor',''),
                           cod_moodle=disc.get('cod_moodle',''), titulacao=disc.get('titulacao',''))
            db.session.add(dd)
        db.session.commit()
        log_action(session['user_id'], session['username'], 'criar', 'course', c.id, c.nome)
        flash('Curso criado com sucesso!', 'success')
        return redirect(url_for('curso_detalhe', id=c.id))
    usuarios = User.query.order_by(User.username).all()
    return render_template('curso_form.html', curso=None, tipos=TIPOS_CURSO, areas=AREAS_VALIDAS, usuarios=usuarios)

@app.route('/cursos/<int:id>')
@login_required
def curso_detalhe(id):
    c    = Course.query.get_or_404(id)
    disc = Discipline.query.filter_by(course_id=id).order_by(Discipline.ordem).all()
    logs = AuditLog.query.filter_by(entity='course', entity_id=id)\
                         .order_by(AuditLog.timestamp.desc()).all()
    extra = json.loads(c.extra_data) if c.extra_data else {}
    return render_template('curso_detalhe.html', c=c, disc=disc, logs=logs, extra=extra)

@app.route('/cursos/<int:id>/editar', methods=['GET','POST'])
@editor_required
def curso_editar(id):
    c = Course.query.get_or_404(id)
    if request.method == 'POST':
        old_nome = c.nome
        d = request.form
        c.nome=d['nome']; c.tipo=d['tipo']; c.area=d.get('area','')
        c.horas=d.get('horas',''); c.meses=d.get('meses',''); c.valor=d.get('valor','')
        c.link_venda=d.get('link_venda',''); c.descricao=d.get('descricao','')
        c.link_imagem=d.get('link_imagem',''); c.obs=d.get('obs','')
        novo_status = d.get('status', c.status)
        # Somente admin pode publicar (ativo); colaboradores podem marcar como finalizado no máximo
        if novo_status == 'ativo' and session.get('role') != 'admin':
            novo_status = c.status
        c.status = novo_status
        c.cupom=d.get('cupom','')
        c.dono=d.get('dono',''); c.ano=d.get('ano',''); c.updated_at=datetime.utcnow()
        ins_list = d.getlist('insersores')
        if ins_list: c.insersor = ','.join(ins_list)
        # Update disciplines — preserva plataforma_ok/plataforma_em por nome da disciplina
        if d.get('disciplinas_json') is not None:
            discs_novas = json.loads(d.get('disciplinas_json','[]') or '[]')
            # Mapa: nome normalizado → disciplina existente (para preservar status da plataforma)
            existentes = {ex.nome.strip().lower(): ex
                         for ex in Discipline.query.filter_by(course_id=id).all()}
            Discipline.query.filter_by(course_id=id).delete()
            for i, disc in enumerate(discs_novas):
                nome = disc.get('nome','')
                anterior = existentes.get(nome.strip().lower())
                dd = Discipline(
                    course_id=id,
                    modulo=disc.get('modulo',''),
                    ordem=i+1,
                    nome=nome,
                    carga=disc.get('carga',''),
                    professor=disc.get('professor',''),
                    cod_moodle=disc.get('cod_moodle',''),
                    titulacao=disc.get('titulacao',''),
                    plataforma_ok=anterior.plataforma_ok if anterior else False,
                    plataforma_em=anterior.plataforma_em if anterior else None
                )
                db.session.add(dd)
        db.session.commit()
        log_action(session['user_id'], session['username'], 'editar', 'course', id, f'{old_nome} → {c.nome}')
        flash('Curso atualizado!', 'success')
        if d.get('from_page') == 'matrizes':
            return redirect(url_for('matrizes'))
        return redirect(url_for('curso_detalhe', id=id))
    disc = Discipline.query.filter_by(course_id=id).order_by(Discipline.ordem).all()
    tipos = ['pos','profissionalizante','rapido','pacote','terceiros','evento','pratica_conectada','pratica_estagio','projeto_ambiental','ggbr','integra_edu']
    usuarios = User.query.order_by(User.username).all()
    return render_template('curso_form.html', curso=c, disc=disc, tipos=TIPOS_CURSO, areas=AREAS_VALIDAS, usuarios=usuarios)

@app.route('/cursos/<int:id>/arquivar', methods=['POST'])
@editor_required
def curso_arquivar(id):
    c = Course.query.get_or_404(id)
    c.status = 'descontinuado'
    db.session.commit()
    log_action(session['user_id'], session['username'], 'arquivar', 'course', id, c.nome)
    flash('Curso arquivado (não excluído).', 'warning')
    return redirect(url_for('cursos'))

# Admins can hard-delete
@app.route('/cursos/<int:id>/excluir', methods=['POST'])
@admin_required
def curso_excluir(id):
    c = Course.query.get_or_404(id)
    nome = c.nome
    Discipline.query.filter_by(course_id=id).delete()
    db.session.delete(c)
    db.session.commit()
    log_action(session['user_id'], session['username'], 'excluir', 'course', id, nome)
    flash(f'Curso "{nome}" excluído permanentemente.', 'danger')
    return redirect(url_for('cursos'))

# ─── API SEARCH ────────────────────────────────────────────────────────────────

@app.route('/api/busca')
@login_required
def api_busca():
    q = request.args.get('q','')
    if len(q) < 2: return jsonify([])
    results = Course.query.filter(Course.nome.ilike(f'%{q}%')).limit(10).all()
    return jsonify([{'id': c.id, 'nome': c.nome, 'tipo': c.tipo, 'status': c.status} for c in results])

# ─── CUPONS ────────────────────────────────────────────────────────────────────

@app.route('/cupons')
@perm_check('can_manage_cupons')
def cupons():
    cupons = Coupon.query.order_by(Coupon.created_at.desc()).all()
    return render_template('cupons.html', cupons=cupons)

@app.route('/cupons/novo', methods=['GET','POST'])
@editor_required
def cupom_novo():
    if request.method == 'POST':
        d = request.form
        def parse_date(s):
            try: return datetime.strptime(s, '%Y-%m-%d').date()
            except: return None
        cp = Coupon(nome=d['nome'], quantidade=int(d.get('quantidade',0) or 0),
                    desconto=float(d.get('desconto',0) or 0),
                    cursos_tipo=d.get('cursos_tipo',''), limite_curso=int(d.get('limite_curso',1) or 1),
                    uso_unico=(d.get('uso_unico')=='SIM'),
                    data_inicial=parse_date(d.get('data_inicial','')),
                    data_final=parse_date(d.get('data_final','')),
                    obs=d.get('obs',''))
        db.session.add(cp)
        db.session.commit()
        log_action(session['user_id'], session['username'], 'criar', 'cupom', cp.id, cp.nome)
        flash('Cupom criado!', 'success')
        return redirect(url_for('cupons'))
    return render_template('cupom_form.html', cupom=None)

@app.route('/cupons/<int:id>/editar', methods=['GET','POST'])
@editor_required
def cupom_editar(id):
    cp = Coupon.query.get_or_404(id)
    if request.method == 'POST':
        d = request.form
        def parse_date(s):
            try: return datetime.strptime(s, '%Y-%m-%d').date()
            except: return None
        cp.nome=d['nome']; cp.quantidade=int(d.get('quantidade',0) or 0)
        cp.desconto=float(d.get('desconto',0) or 0)
        cp.cursos_tipo=d.get('cursos_tipo','')
        cp.limite_curso=int(d.get('limite_curso',1) or 1)
        cp.uso_unico=(d.get('uso_unico')=='SIM')
        cp.data_inicial=parse_date(d.get('data_inicial',''))
        cp.data_final=parse_date(d.get('data_final',''))
        cp.obs=d.get('obs','')
        db.session.commit()
        log_action(session['user_id'], session['username'], 'editar', 'cupom', id, cp.nome)
        flash('Cupom atualizado!', 'success')
        return redirect(url_for('cupons'))
    return render_template('cupom_form.html', cupom=cp)

# ─── REEMBOLSOS ────────────────────────────────────────────────────────────────

@app.route('/reembolsos')
@perm_check('can_manage_reembolsos')
def reembolsos():
    busca      = request.args.get('q', '').strip()
    f_colab    = request.args.get('colab', '').strip()
    f_cat      = request.args.get('categoria', '').strip()
    f_pend     = request.args.get('pendencia', '').strip()

    q = Refund.query
    if busca:
        q = q.filter(db.or_(
            Refund.nome_aluno.ilike(f'%{busca}%'),
            Refund.nome_curso.ilike(f'%{busca}%')
        ))
    if f_colab:
        q = q.filter(Refund.colab.ilike(f'%{f_colab}%'))
    if f_cat:
        q = q.filter(Refund.categoria.ilike(f'%{f_cat}%'))

    items = q.order_by(Refund.created_at.desc()).all()

    # Filtro por pendência (feito em Python pois usa property)
    if f_pend:
        items = [r for r in items if r.pendencia[0] == f_pend]

    contagem = {'sem_solic1': 0, 'sem_solic2': 0, 'sem_aprovacao': 0, 'sem_exclusao': 0, 'concluido': 0}
    for r in Refund.query.all():
        contagem[r.pendencia[0]] += 1

    colabs = sorted({r.colab for r in Refund.query.all() if r.colab})
    cats   = sorted({r.categoria for r in Refund.query.all() if r.categoria})

    return render_template('reembolsos.html', items=items, contagem=contagem,
                           busca=busca, f_colab=f_colab, f_cat=f_cat, f_pend=f_pend,
                           colabs=colabs, cats=cats)

@app.route('/reembolsos/novo', methods=['GET','POST'])
@editor_required
def reembolso_novo():
    if request.method == 'POST':
        r = _refund_from_form(request.form)
        db.session.add(r)
        db.session.commit()
        log_action(session['user_id'], session['username'], 'criar', 'reembolso', r.id, r.nome_aluno)
        flash('Reembolso registrado!', 'success')
        return redirect(url_for('reembolsos'))
    return render_template('reembolso_form.html', item=None)

@app.route('/reembolsos/<int:id>/editar', methods=['GET','POST'])
@editor_required
def reembolso_editar(id):
    r = Refund.query.get_or_404(id)
    if request.method == 'POST':
        d = request.form
        def pdate(s):
            try: return datetime.strptime(s, '%Y-%m-%d').date()
            except: return None
        r.colab          = d.get('colab','')
        r.nome_aluno     = d.get('nome_aluno','')
        r.nome_curso     = d.get('nome_curso','')
        r.categoria      = d.get('categoria','')
        r.data_compra    = pdate(d.get('data_compra',''))
        r.data_solicitacao = pdate(d.get('data_solicitacao',''))
        r.valor          = float(d.get('valor',0) or 0)
        r.valor_estorno  = float(d.get('valor_estorno',0) or 0)
        r.solicitacao_1  = pdate(d.get('solicitacao_1',''))
        r.solicitacao_2  = pdate(d.get('solicitacao_2',''))
        r.data_aprovacao = pdate(d.get('data_aprovacao',''))
        r.motivo           = d.get('motivo','')
        r.curso_excluido   = pdate(d.get('curso_excluido',''))
        r.obs              = d.get('obs','')
        r.concluido_manual = d.get('concluido_manual') == 'on'
        r.cpf              = d.get('cpf','')
        r.celular          = d.get('celular','')
        r.pix              = d.get('pix','')
        r.email_destino    = d.get('email_destino','')
        db.session.commit()
        log_action(session['user_id'], session['username'], 'editar', 'reembolso', id, r.nome_aluno)
        flash('Reembolso atualizado!', 'success')
        return redirect(url_for('reembolsos'))
    return render_template('reembolso_form.html', item=r)

@app.route('/reembolsos/<int:id>/enviar-email', methods=['POST'])
@login_required
def reembolso_enviar_email(id):
    r = Refund.query.get_or_404(id)
    if not r.email_destino:
        flash('Preencha o campo "E-mail de Destino" e salve antes de enviar.', 'danger')
        return redirect(url_for('reembolso_editar', id=id))
    valor_pago = ('R$ ' + f'{r.valor:.2f}'.replace('.', ',')) if r.valor else 'R$ —'
    valor_rec_num = r.valor_estorno or r.valor
    valor_rec = ('R$ ' + f'{valor_rec_num:.2f}'.replace('.', ',')) if valor_rec_num else 'R$ —'
    assunto = f'SOLICITAÇÃO DE REEMBOLSO INOVA CARREIRA - {r.nome_aluno}'
    corpo = (
        'Prezado(a)s,\n\n'
        'Por favor, solicito o pagamento de reembolso para o(a) seguinte aluno(a) matriculado(a) '
        f'no curso {r.nome_curso}, {r.motivo or ""}, encaminhado em anexo comprovantes de '
        'pagamento do aluno e do sistema da plataforma.\n\n'
        f'Valor Pago = {valor_pago}\n'
        f'Valor a receber: {valor_rec}\n\n'
        'Dados para pagamento:\n\n'
        f'Nome: {r.nome_aluno}\n'
        f'CPF : {r.cpf or "[CPF]"}\n'
        f'Celular: {r.celular or "[CELULAR]"}\n\n'
        f'PIX: {r.pix or "[PIX]"}\n\n\n'
        'Atenciosamente,\nINOVA Carreira'
    )
    if enviar_email(r.email_destino, assunto, corpo):
        log_action(session['user_id'], session['username'], 'enviar_email', 'reembolso', r.id, r.nome_aluno)
        flash(f'E-mail enviado para {r.email_destino}!', 'success')
    else:
        flash('Não foi possível enviar o e-mail agora. Tente novamente em instantes.', 'danger')
    return redirect(url_for('reembolso_editar', id=id))

@app.route('/reembolsos/<int:id>/excluir', methods=['POST'])
@admin_required
def reembolso_excluir(id):
    r = Refund.query.get_or_404(id)
    nome = r.nome_aluno
    db.session.delete(r)
    db.session.commit()
    log_action(session['user_id'], session['username'], 'excluir', 'reembolso', id, nome)
    flash(f'Reembolso de "{nome}" excluído.', 'success')
    return redirect(url_for('reembolsos'))

@app.route('/reembolsos/marcar-todos-concluidos', methods=['POST'])
@admin_required
def reembolsos_marcar_todos_concluidos():
    total = Refund.query.filter_by(concluido_manual=False).update({'concluido_manual': True})
    db.session.commit()
    log_action(session['user_id'], session['username'], 'marcar_todos', 'reembolso', None, f'{total} reembolsos')
    flash(f'{total} reembolso(s) marcado(s) como concluído.', 'success')
    return redirect(url_for('reembolsos'))

def _refund_from_form(d):
    def pdate(s):
        try: return datetime.strptime(s, '%Y-%m-%d').date()
        except: return None
    return Refund(
        colab=d.get('colab',''), nome_aluno=d.get('nome_aluno',''),
        nome_curso=d.get('nome_curso',''), categoria=d.get('categoria',''),
        data_compra=pdate(d.get('data_compra','')),
        data_solicitacao=pdate(d.get('data_solicitacao','')),
        valor=float(d.get('valor',0) or 0),
        valor_estorno=float(d.get('valor_estorno',0) or 0),
        solicitacao_1=pdate(d.get('solicitacao_1','')),
        solicitacao_2=pdate(d.get('solicitacao_2','')),
        data_aprovacao=pdate(d.get('data_aprovacao','')),
        motivo=d.get('motivo',''),
        curso_excluido=pdate(d.get('curso_excluido','')),
        obs=d.get('obs',''),
        concluido_manual=d.get('concluido_manual') == 'on',
        cpf=d.get('cpf',''),
        celular=d.get('celular',''),
        pix=d.get('pix',''),
        email_destino=d.get('email_destino',''),
    )

# ─── MATRIZES CURRICULARES ─────────────────────────────────────────────────────

@app.route('/disciplina/<int:disc_id>/toggle', methods=['POST'])
@login_required
def disciplina_toggle(disc_id):
    d = Discipline.query.get_or_404(disc_id)
    d.plataforma_ok = not d.plataforma_ok
    d.plataforma_em = datetime.utcnow() if d.plataforma_ok else None
    db.session.commit()
    return jsonify({'ok': d.plataforma_ok, 'disc_id': disc_id})

@app.route('/curso/<int:course_id>/disciplinas/marcar-todas', methods=['POST'])
@login_required
def disciplinas_marcar_todas(course_id):
    marcar = request.json.get('marcar', True)
    discs = Discipline.query.filter_by(course_id=course_id).all()
    now = datetime.utcnow()
    for d in discs:
        d.plataforma_ok = marcar
        d.plataforma_em = now if marcar else None
    db.session.commit()
    return jsonify({'ok': True, 'total': len(discs), 'marcar': marcar})

@app.route('/banco-disciplinas')
@login_required
def banco_disciplinas():
    busca = request.args.get('q', '').strip()

    import unicodedata, re as _re
    def _norm(s):
        s = _re.sub(r'\s+', ' ', s.upper().strip())
        s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
        return s

    todas = Discipline.query.order_by(Discipline.nome).all()
    cursos_map = {c.id: c for c in Course.query.all()}

    grupos = {}
    for d in todas:
        chave = _norm(d.nome)
        if busca and busca.lower() not in d.nome.lower():
            continue
        if chave not in grupos:
            grupos[chave] = {'nome': d.nome, 'cursos': [], 'ocorrencias': []}
        curso = cursos_map.get(d.course_id)
        if curso and curso.id not in [c.id for c in grupos[chave]['cursos']]:
            grupos[chave]['cursos'].append(curso)
        grupos[chave]['ocorrencias'].append(d)

    disciplinas_unicas = sorted(grupos.values(), key=lambda x: x['nome'])
    return render_template('banco_disciplinas.html',
                           disciplinas=disciplinas_unicas, busca=busca,
                           total_unicas=len(disciplinas_unicas),
                           total_ocorrencias=len(todas))


@app.route('/banco-disciplinas/exportar-excel')
@login_required
def banco_disciplinas_exportar_excel():
    import openpyxl, re as _re, unicodedata
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    busca = request.args.get('q', '').strip()

    def _norm(s):
        s = _re.sub(r'\s+', ' ', s.upper().strip())
        return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

    TIPO_LABEL = {
        'pos': 'Pós-Graduação', 'profissionalizante': 'Profissionalizante',
        'rapido': 'Rápido', 'pacote': 'Pacote', 'terceiros': 'Terceiros',
        'evento': 'Evento', 'pratica_conectada': 'Prática Conectada',
        'pratica_estagio': 'Prática Estágio', 'projeto_ambiental': 'Proj. Ambiental',
        'ggbr': 'GGBR', 'integra_edu': 'Integra Edu',
    }

    todas = Discipline.query.order_by(Discipline.nome).all()
    cursos_map = {c.id: c for c in Course.query.all()}
    grupos = {}
    for d in todas:
        if busca and busca.lower() not in d.nome.lower():
            continue
        chave = _norm(d.nome)
        if chave not in grupos:
            grupos[chave] = {'nome': d.nome, 'ocorrencias': []}
        grupos[chave]['ocorrencias'].append((d, cursos_map.get(d.course_id)))

    wb = openpyxl.Workbook()

    # ── ABA 1: DISCIPLINAS × CURSOS (agrupado) ──────────────────────────────
    ws1 = wb.active
    ws1.title = 'Banco de Disciplinas'

    thin = Side(style='thin', color='E8E2DA')
    med  = Side(style='medium', color='F97316')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_disc = Border(left=med, right=thin, top=thin, bottom=thin)

    # Título
    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value = f'Gestor Acadêmico — Banco de Disciplinas'
    title_cell.font = Font(bold=True, size=14, color='F97316')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    title_cell.fill = PatternFill('solid', fgColor='FFF7ED')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:H2')
    sub_cell = ws1['A2']
    sub_cell.value = f'Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")} · {len(grupos)} disciplinas únicas · {len(todas)} ocorrências'
    sub_cell.font = Font(size=9, color='78716C', italic=True)
    sub_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[2].height = 18

    # Cabeçalho
    hfill  = PatternFill('solid', fgColor='F97316')
    hfont  = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap   = Alignment(wrap_text=True, vertical='top')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = ['Disciplina', 'CH', 'Módulo', 'Curso', 'Tipo', 'Área', 'Professor', 'Titulação']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=3, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr
    ws1.row_dimensions[3].height = 24
    ws1.freeze_panes = 'A4'

    fill_disc  = PatternFill('solid', fgColor='FFF7ED')  # linha de disciplina (laranja suave)
    fill_curso = PatternFill('solid', fgColor='FFFFFF')   # linha de curso (branco)
    fill_alt   = PatternFill('solid', fgColor='FAFAF9')   # linha alternada

    row_idx = 4
    for idx, chave in enumerate(sorted(grupos.keys())):
        grupo = grupos[chave]
        ocors = grupo['ocorrencias']
        n_ocors = len(ocors)

        for i, (d, curso) in enumerate(ocors):
            is_first = (i == 0)
            tipo_label = TIPO_LABEL.get(curso.tipo, curso.tipo) if curso else ''
            fill = fill_disc if is_first else (fill_curso if i % 2 == 0 else fill_alt)

            vals = [
                grupo['nome'] if is_first else '',
                d.carga or '',
                d.modulo or '',
                curso.nome if curso else '',
                tipo_label,
                curso.area or '' if curso else '',
                d.professor or '',
                d.titulacao or '',
            ]
            for col, val in enumerate(vals, 1):
                cell = ws1.cell(row=row_idx, column=col, value=val)
                cell.border = bdr_disc if col == 1 else bdr
                cell.alignment = wrap
                cell.fill = fill
                if col == 1 and is_first:
                    cell.font = Font(bold=True, size=10)
            ws1.row_dimensions[row_idx].height = 16
            row_idx += 1

        # Linha separadora entre disciplinas
        if idx < len(grupos) - 1:
            for col in range(1, 9):
                cell = ws1.cell(row=row_idx, column=col, value='')
                cell.fill = PatternFill('solid', fgColor='F97316')
                cell.border = Border(top=Side(style='hair', color='F97316'))
            ws1.row_dimensions[row_idx].height = 3
            row_idx += 1

    for i, w in enumerate([44, 6, 16, 52, 18, 14, 24, 16], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── ABA 2: LISTA COMPLETA (uma linha por ocorrência) ────────────────────
    ws2 = wb.create_sheet('Lista Completa')

    ws2.merge_cells('A1:H1')
    t2 = ws2['A1']
    t2.value = 'Gestor Acadêmico — Lista Completa de Disciplinas'
    t2.font = Font(bold=True, size=13, color='F97316')
    t2.alignment = Alignment(horizontal='left', vertical='center')
    t2.fill = PatternFill('solid', fgColor='FFF7ED')
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center; c.border = bdr
    ws2.row_dimensions[2].height = 22
    ws2.freeze_panes = 'A3'

    row2 = 3
    for chave in sorted(grupos.keys()):
        for d, curso in grupos[chave]['ocorrencias']:
            tipo_label = TIPO_LABEL.get(curso.tipo, curso.tipo) if curso else ''
            vals = [
                grupos[chave]['nome'], d.carga or '', d.modulo or '',
                curso.nome if curso else '',
                tipo_label,
                curso.area or '' if curso else '',
                d.professor or '', d.titulacao or '',
            ]
            fill_row = PatternFill('solid', fgColor='FFFFFF') if row2 % 2 == 1 else PatternFill('solid', fgColor='FAFAF9')
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=row2, column=col, value=val)
                cell.border = bdr; cell.alignment = wrap; cell.fill = fill_row
            ws2.row_dimensions[row2].height = 15
            row2 += 1

    for i, w in enumerate([44, 6, 16, 52, 18, 14, 24, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── ABA 3: RESUMO ESTATÍSTICO ────────────────────────────────────────────
    ws3 = wb.create_sheet('Resumo')
    ws3.merge_cells('A1:C1')
    t3 = ws3['A1']
    t3.value = 'Gestor Acadêmico — Resumo do Banco de Disciplinas'
    t3.font = Font(bold=True, size=13, color='F97316')
    t3.alignment = Alignment(horizontal='left', vertical='center')
    t3.fill = PatternFill('solid', fgColor='FFF7ED')
    ws3.row_dimensions[1].height = 28

    resumo_dados = [
        ('Total de disciplinas únicas', len(grupos)),
        ('Total de ocorrências', len(todas)),
        ('Média de cursos por disciplina', f'{len(todas)/len(grupos):.1f}' if grupos else '0'),
        ('', ''),
        ('Disciplinas com mais ocorrências', ''),
    ]
    row3 = 2
    for label, val in resumo_dados:
        ws3.cell(row=row3, column=1, value=label).font = Font(bold=True if not val else False, size=10)
        ws3.cell(row=row3, column=2, value=val).font = Font(bold=True, color='F97316', size=11)
        row3 += 1

    top_discs = sorted(grupos.values(), key=lambda x: len(x['ocorrencias']), reverse=True)[:10]
    for item in top_discs:
        ws3.cell(row=row3, column=1, value=item['nome']).font = Font(size=10)
        ws3.cell(row=row3, column=2, value=f"{len(item['ocorrencias'])} curso(s)").font = Font(color='F97316')
        row3 += 1

    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    suffix = f'_{busca}' if busca else ''
    fname = f'banco_disciplinas{suffix}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@app.route('/banco-disciplinas/relatorio')
@login_required
def banco_disciplinas_relatorio():
    busca = request.args.get('q', '').strip()

    import unicodedata, re as _re
    def _norm(s):
        s = _re.sub(r'\s+', ' ', s.upper().strip())
        return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

    TIPO_LABEL = {
        'pos': 'Pós-Graduação', 'profissionalizante': 'Profissionalizante',
        'rapido': 'Rápido', 'pacote': 'Pacote', 'terceiros': 'Terceiros',
        'evento': 'Evento', 'pratica_conectada': 'Prática Conectada',
        'pratica_estagio': 'Prática Estágio', 'projeto_ambiental': 'Proj. Ambiental',
        'ggbr': 'GGBR', 'integra_edu': 'Integra Edu',
    }

    todas = Discipline.query.order_by(Discipline.nome).all()
    cursos_map = {c.id: c for c in Course.query.all()}
    grupos = {}
    for d in todas:
        if busca and busca.lower() not in d.nome.lower():
            continue
        chave = _norm(d.nome)
        if chave not in grupos:
            grupos[chave] = {'nome': d.nome, 'cursos': [], 'ocorrencias': []}
        curso = cursos_map.get(d.course_id)
        if curso and curso.id not in [c.id for c in grupos[chave]['cursos']]:
            grupos[chave]['cursos'].append(curso)
        grupos[chave]['ocorrencias'].append((d, curso))

    disciplinas = sorted(grupos.values(), key=lambda x: x['nome'])
    total_ocors = len(todas)
    return render_template('banco_disciplinas_relatorio.html',
                           disciplinas=disciplinas, busca=busca,
                           total_unicas=len(disciplinas),
                           total_ocorrencias=total_ocors,
                           tipo_label=TIPO_LABEL,
                           gerado_em=datetime.now())


@app.route('/ia-assistente')
@login_required
def ia_assistente():
    cursos_amostra = Course.query.filter(Course.status != 'descontinuado').order_by(Course.nome).limit(20).all()
    return render_template('ia_assistente.html', cursos_amostra=cursos_amostra)


@app.route('/api/ia/chat', methods=['POST'])
@login_required
def ia_chat():
    import unicodedata as _ud, re as _re

    data = request.json or {}
    pergunta = data.get('pergunta', '').strip()
    if not pergunta:
        return jsonify({'erro': 'Pergunta vazia'}), 400

    def _norm(s):
        s = _re.sub(r'\s+', ' ', s.upper().strip())
        return ''.join(c for c in _ud.normalize('NFKD', s) if not _ud.combining(c))

    def _contem(texto, *palavras):
        t = _norm(texto)
        return any(p in t for p in [_norm(w) for w in palavras])

    def _tipo_label(tipo):
        return {
            'pos': 'Pós-Graduação', 'profissionalizante': 'Profissionalizante',
            'rapido': 'Curso Rápido', 'pacote': 'Pacote', 'terceiros': 'Terceiros',
            'evento': 'Evento', 'pratica_conectada': 'Prática Conectada',
            'pratica_estagio': 'Prática Estágio', 'projeto_ambiental': 'Projeto Ambiental',
            'ggbr': 'GGBR', 'integra_edu': 'Integra Edu',
        }.get(tipo, tipo)

    p = pergunta
    linhas = []

    # ── ESTATÍSTICAS GERAIS ─────────────────────────────
    if _contem(p, 'quantos', 'total', 'quantidade', 'estatistica', 'estatística', 'resumo', 'geral'):
        total = Course.query.count()
        ativos = Course.query.filter_by(status='ativo').count()
        em_ed = Course.query.filter_by(status='em_edicao').count()
        desc = Course.query.filter_by(status='descontinuado').count()
        n_disc = Discipline.query.count()
        import unicodedata as _ud2, re as _re2
        def _n2(s):
            s = _re2.sub(r'\s+', ' ', s.upper().strip())
            return ''.join(c for c in _ud2.normalize('NFKD', s) if not _ud2.combining(c))
        todas_d = Discipline.query.with_entities(Discipline.nome).all()
        unicas = len({_n2(d.nome) for d in todas_d})
        linhas = [
            f"Aqui está o resumo geral do sistema Gestor Acadêmico:\n",
            f"**Cursos cadastrados:** {total}",
            f"  • Ativos: {ativos}",
            f"  • Em edição: {em_ed}",
            f"  • Descontinuados: {desc}",
            f"\n**Banco de Disciplinas:** {unicas} disciplinas únicas ({n_disc} ocorrências no total)",
        ]
        por_tipo = db.session.query(Course.tipo, db.func.count(Course.id)).group_by(Course.tipo).order_by(db.func.count(Course.id).desc()).all()
        linhas.append("\n**Distribuição por tipo:**")
        for tipo, cnt in por_tipo:
            linhas.append(f"  • {_tipo_label(tipo)}: {cnt} curso(s)")
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── EVENTOS ─────────────────────────────────────────
    if _contem(p, 'evento', 'eventos', 'acontecendo', 'acontece', 'agenda', 'programacao', 'programação'):
        eventos = Course.query.filter_by(tipo='evento').order_by(Course.nome).all()
        if not eventos:
            return jsonify({'ok': True, 'resposta': 'Não há nenhum evento cadastrado no sistema no momento.'})
        linhas = [f"**Eventos cadastrados no sistema ({len(eventos)}):**\n"]
        for e in eventos:
            info = f"• **{e.nome}**"
            if e.area: info += f" — Área: {e.area}"
            if e.horas: info += f" | {e.horas}"
            if e.valor and e.valor not in ['-', '', 'None']: info += f" | R$ {e.valor}"
            if e.status: info += f" | Status: {e.status.replace('_', ' ').title()}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── PACOTES ─────────────────────────────────────────
    if _contem(p, 'pacote', 'pacotes', 'combo', 'bundle'):
        pacotes = Course.query.filter_by(tipo='pacote').order_by(Course.nome).all()
        if not pacotes:
            return jsonify({'ok': True, 'resposta': 'Não há pacotes cadastrados no sistema.'})
        linhas = [f"**Pacotes disponíveis ({len(pacotes)}):**\n"]
        for pac in pacotes:
            n_discs = Discipline.query.filter_by(course_id=pac.id).count()
            info = f"• **{pac.nome}**"
            if pac.valor and pac.valor not in ['-', '', 'None']: info += f" — R$ {pac.valor}"
            if n_discs: info += f" | {n_discs} disciplina(s)/curso(s)"
            if pac.status: info += f" | {pac.status.replace('_', ' ').title()}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── PÓS-GRADUAÇÃO ────────────────────────────────────
    if _contem(p, 'pos', 'pós', 'pos-graduacao', 'pós-graduação', 'graduacao', 'graduação', 'mba', 'especializacao', 'especialização'):
        cursos = Course.query.filter_by(tipo='pos').filter(Course.status != 'descontinuado').order_by(Course.nome).all()
        if not cursos:
            return jsonify({'ok': True, 'resposta': 'Não há cursos de pós-graduação ativos cadastrados.'})
        linhas = [f"**Cursos de Pós-Graduação ({len(cursos)}):**\n"]
        for c in cursos:
            info = f"• **{c.nome}**"
            if c.area: info += f" — {c.area}"
            if c.horas: info += f" | {c.horas}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── PROFISSIONALIZANTES ──────────────────────────────
    if _contem(p, 'profissionalizante', 'profissionalizantes', 'tecnico', 'técnico'):
        cursos = Course.query.filter_by(tipo='profissionalizante').filter(Course.status != 'descontinuado').order_by(Course.nome).all()
        if not cursos:
            return jsonify({'ok': True, 'resposta': 'Não há cursos profissionalizantes ativos.'})
        linhas = [f"**Cursos Profissionalizantes ({len(cursos)}):**\n"]
        for c in cursos:
            info = f"• **{c.nome}**"
            if c.area: info += f" — {c.area}"
            if c.horas: info += f" | {c.horas}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── CURSOS RÁPIDOS ───────────────────────────────────
    if _contem(p, 'rapido', 'rápido', 'rapidos', 'rápidos', 'curto', 'curta duracao', 'curta duração'):
        cursos = Course.query.filter_by(tipo='rapido').filter(Course.status != 'descontinuado').order_by(Course.nome).all()
        linhas = [f"**Cursos Rápidos ({len(cursos)}):**\n"]
        for c in cursos:
            info = f"• **{c.nome}**"
            if c.area: info += f" — {c.area}"
            if c.horas: info += f" | {c.horas}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas) if linhas else 'Nenhum curso rápido cadastrado.'})

    # ── BUSCA POR ÁREA ───────────────────────────────────
    areas_map = {
        'saude': 'SAÚDE', 'saúde': 'SAÚDE',
        'negocios': 'NEGÓCIOS', 'negócios': 'NEGÓCIOS', 'negocio': 'NEGÓCIOS',
        'tecnologia': 'TECNOLOGIA', 'ti': 'TECNOLOGIA', 'informatica': 'TECNOLOGIA',
        'educacao': 'EDUCAÇÃO', 'educação': 'EDUCAÇÃO', 'pedagogia': 'EDUCAÇÃO',
        'criatividade': 'CRIATIVIDADE', 'design': 'CRIATIVIDADE', 'arte': 'CRIATIVIDADE',
        'gastronomia': 'GASTRONOMIA', 'culinaria': 'GASTRONOMIA', 'culinária': 'GASTRONOMIA',
    }
    area_encontrada = None
    p_norm = _norm(p)
    for chave, area_val in areas_map.items():
        if _norm(chave) in p_norm:
            area_encontrada = area_val
            break
    if area_encontrada:
        cursos = Course.query.filter_by(area=area_encontrada).filter(Course.status != 'descontinuado').order_by(Course.nome).all()
        if not cursos:
            return jsonify({'ok': True, 'resposta': f'Não encontrei cursos ativos na área de **{area_encontrada}**.'})
        linhas = [f"**Cursos na área de {area_encontrada} ({len(cursos)}):**\n"]
        for c in cursos:
            info = f"• **{c.nome}** ({_tipo_label(c.tipo)})"
            if c.horas: info += f" | {c.horas}"
            linhas.append(info)
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── DISCIPLINAS DO BANCO ─────────────────────────────
    if _contem(p, 'banco de disciplinas', 'disciplinas unicas', 'disciplinas únicas', 'todas as disciplinas', 'quais disciplinas temos', 'disciplinas disponiveis', 'disciplinas disponíveis'):
        todas = Discipline.query.with_entities(Discipline.nome).all()
        unicas = sorted({_norm(d.nome): d.nome for d in todas}.values())
        linhas = [f"**Banco de Disciplinas — {len(unicas)} disciplinas únicas cadastradas:**\n"]
        for i, nome in enumerate(unicas, 1):
            linhas.append(f"{i}. {nome}")
        return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── DISCIPLINAS PARA UM CURSO / SUGESTÃO ─────────────
    # Detecta padrões: "disciplinas para X", "disciplinas de X", "encaixam para X", "grade de X", "matriz de X"
    padroes_curso = [
        r'disciplinas?\s+(?:para|de|do|da)\s+(?:curso\s+(?:de\s+)?)?([\w\s]+)',
        r'encaixam?\s+(?:para|em|no|na)\s+(?:curso\s+(?:de\s+)?)?([\w\s]+)',
        r'grade\s+(?:curricular\s+)?(?:do|da|de)?\s+([\w\s]+)',
        r'matriz\s+(?:do|da|de)?\s+([\w\s]+)',
        r'sugest[aã]o\s+(?:para|de)?\s+(?:curso\s+(?:de\s+)?)?([\w\s]+)',
        r'sugira\s+(?:disciplinas?\s+)?(?:para|de)?\s+(?:curso\s+(?:de\s+)?)?([\w\s]+)',
    ]
    tema_busca = None
    for padrao in padroes_curso:
        m = _re.search(padrao, p, _re.IGNORECASE)
        if m:
            tema_busca = m.group(1).strip().rstrip('?.,! ')
            break

    # Também detecta perguntas sem padrão estruturado, ex: "o que tem no curso de administração"
    if not tema_busca:
        m = _re.search(r'curso\s+(?:de\s+|do\s+|da\s+)?([\w\s]{3,40}?)(?:\?|$|,|\.)', p, _re.IGNORECASE)
        if m:
            tema_busca = m.group(1).strip()

    if tema_busca:
        tema_norm = _norm(tema_busca)
        # 1. Busca curso exato ou similar no banco
        todos_cursos = Course.query.order_by(Course.nome).all()
        cursos_match = [c for c in todos_cursos if tema_norm in _norm(c.nome) or _norm(c.nome) in tema_norm]

        if cursos_match:
            # Encontrou curso(s) correspondente(s) — mostra as disciplinas
            linhas = []
            for c in cursos_match[:3]:
                discs = Discipline.query.filter_by(course_id=c.id).order_by(Discipline.ordem).all()
                linhas.append(f"**{c.nome}** ({_tipo_label(c.tipo)})")
                if c.area: linhas.append(f"Área: {c.area}")
                if discs:
                    linhas.append(f"Disciplinas da matriz ({len(discs)}):\n")
                    for d in discs:
                        item = f"• {d.nome}"
                        if d.carga: item += f" — {d.carga}"
                        if d.modulo: item += f" | Módulo: {d.modulo}"
                        linhas.append(item)
                else:
                    linhas.append("_(Este curso ainda não tem matriz cadastrada)_")
                linhas.append('')
            return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

        else:
            # Não achou curso — sugere disciplinas do banco que se relacionam ao tema
            todas_disc = Discipline.query.order_by(Discipline.nome).all()
            cursos_map = {c.id: c for c in todos_cursos}

            # Busca disciplinas cujo nome tenha palavras em comum com o tema
            palavras_tema = set(tema_norm.split())
            palavras_tema -= {'DE', 'DA', 'DO', 'EM', 'E', 'O', 'A', 'PARA', 'COM'}

            grupos = {}
            for d in todas_disc:
                chave = _norm(d.nome)
                if chave not in grupos:
                    grupos[chave] = {'nome': d.nome, 'cursos': [], 'relevancia': 0}
                # Calcula relevância por palavras em comum
                palavras_disc = set(_norm(d.nome).split())
                comuns = palavras_tema & palavras_disc
                grupos[chave]['relevancia'] = max(grupos[chave]['relevancia'], len(comuns))
                curso = cursos_map.get(d.course_id)
                if curso and curso.nome not in grupos[chave]['cursos']:
                    grupos[chave]['cursos'].append(curso.nome)

            relevantes = sorted(
                [v for v in grupos.values() if v['relevancia'] > 0],
                key=lambda x: -x['relevancia']
            )[:15]

            if relevantes:
                linhas = [
                    f"Não encontrei um curso com o nome **\"{tema_busca}\"** no sistema.",
                    f"Mas encontrei **{len(relevantes)} disciplinas** do nosso banco que podem se encaixar:\n"
                ]
                for item in relevantes:
                    linha = f"• **{item['nome']}**"
                    if item['cursos']:
                        linha += f" — presente em: {', '.join(item['cursos'][:2])}"
                        if len(item['cursos']) > 2: linha += f" +{len(item['cursos'])-2}"
                    linhas.append(linha)
                linhas.append(f"\n💡 Você pode ver o banco completo em **Banco de Disciplinas** no menu lateral.")
            else:
                # Busca mais ampla: qualquer disciplina, lista as mais usadas
                from sqlalchemy import func as sqlfunc
                top_discs = db.session.query(
                    Discipline.nome, sqlfunc.count(Discipline.id).label('cnt')
                ).group_by(Discipline.nome).order_by(sqlfunc.count(Discipline.id).desc()).limit(20).all()

                linhas = [
                    f"Não encontrei correspondências diretas para **\"{tema_busca}\"**.",
                    f"Aqui estão as disciplinas mais utilizadas nos nossos cursos que podem servir de base:\n"
                ]
                for nome, cnt in top_discs:
                    linhas.append(f"• {nome} ({cnt} curso(s))")
                linhas.append(f"\n💡 Acesse **Banco de Disciplinas** para ver todas as {len(grupos)} disciplinas únicas.")

            return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── BUSCA LIVRE POR NOME DE CURSO ────────────────────
    palavras = [w for w in _norm(p).split() if len(w) > 3 and w not in {'COMO', 'QUAL', 'QUAIS', 'ONDE', 'QUANDO', 'QUERO', 'TENHO', 'TEMOS', 'ESTA', 'ESTAO', 'SOBRE', 'MOSTRAR', 'LISTAR', 'LISTA', 'PODE', 'EXISTEM', 'EXISTE', 'CADASTRADO', 'SISTEMA'}]
    if palavras:
        resultados = []
        todos = Course.query.filter(Course.status != 'descontinuado').all()
        for c in todos:
            cnorm = _norm(c.nome)
            if any(w in cnorm for w in palavras):
                resultados.append(c)
        if resultados:
            linhas = [f"**Encontrei {len(resultados)} curso(s) relacionado(s):**\n"]
            for c in resultados[:15]:
                info = f"• **{c.nome}** ({_tipo_label(c.tipo)})"
                if c.area: info += f" — {c.area}"
                if c.horas: info += f" | {c.horas}"
                if c.status: info += f" | {c.status.replace('_',' ').title()}"
                linhas.append(info)
            if len(resultados) > 15:
                linhas.append(f"\n_...e mais {len(resultados)-15} curso(s). Refine a busca para ver todos._")
            return jsonify({'ok': True, 'resposta': '\n'.join(linhas)})

    # ── RESPOSTA PADRÃO ──────────────────────────────────
    total = Course.query.filter(Course.status != 'descontinuado').count()
    import unicodedata as _ud3, re as _re3
    def _n3(s):
        s = _re3.sub(r'\s+', ' ', s.upper().strip())
        return ''.join(c for c in _ud3.normalize('NFKD', s) if not _ud3.combining(c))
    unicas = len({_n3(d.nome) for d in Discipline.query.with_entities(Discipline.nome).all()})

    resposta = (
        f"Posso te ajudar a encontrar informações no sistema Gestor Acadêmico. "
        f"Temos **{total} cursos ativos** e **{unicas} disciplinas** no banco.\n\n"
        f"Experimente me perguntar:\n"
        f"• _\"Quais eventos estão cadastrados?\"_\n"
        f"• _\"Disciplinas para o curso de Administração\"_\n"
        f"• _\"Liste os cursos de pós-graduação\"_\n"
        f"• _\"Cursos na área de Saúde\"_\n"
        f"• _\"Quais pacotes temos?\"_\n"
        f"• _\"Resumo geral do sistema\"_"
    )
    return jsonify({'ok': True, 'resposta': resposta})


@app.route('/admin/marcar-tudo-concluido', methods=['POST'])
@admin_required
def admin_marcar_concluido():
    now = datetime.utcnow()
    total = Discipline.query.filter_by(plataforma_ok=False).update(
        {'plataforma_ok': True, 'plataforma_em': now}
    )
    db.session.commit()
    log_action(session['user_id'], session['username'], 'marcar_tudo', 'discipline', None,
               f'Marcou {total} disciplinas como concluídas')
    flash(f'{total} disciplina(s) marcada(s) como concluída(s)!', 'success')
    return redirect(request.referrer or url_for('matrizes'))


@app.route('/pacotes')
@login_required
def pacotes():
    busca = request.args.get('q', '').strip()

    todos = Course.query.filter_by(tipo='pacote').order_by(Course.nome).all()

    if busca:
        busca_low = busca.lower()
        disc_cids = {r[0] for r in db.session.query(Discipline.course_id)
                     .filter(Discipline.nome.ilike(f'%{busca}%')).all()}
        todos = [c for c in todos if busca_low in c.nome.lower() or c.id in disc_cids]

    pacote_data = []
    for c in todos:
        discs = Discipline.query.filter_by(course_id=c.id).order_by(Discipline.ordem).all()
        if busca and busca.lower() not in c.nome.lower():
            discs = [d for d in discs if busca.lower() in d.nome.lower()]
        pacote_data.append({'course': c, 'disciplines': discs})

    return render_template('pacotes.html', pacote_data=pacote_data, busca=busca)


@app.route('/admin/migrar-externos-para-pacotes', methods=['POST'])
@admin_required
def admin_migrar_externos():
    externos = Course.query.filter_by(status='externo').all()
    count = len(externos)
    for c in externos:
        c.tipo = 'pacote'
        c.status = 'ativo'
    db.session.commit()
    log_action(session['user_id'], session['username'], 'migrar', 'course', None,
               f'Migrou {count} curso(s) de status=externo para tipo=pacote')
    flash(f'{count} curso(s) migrado(s) de "Externo" para "Pacote" com sucesso!', 'success')
    return redirect(url_for('pacotes'))


@app.route('/matrizes')
@login_required
def matrizes():
    busca = request.args.get('q', '').strip()
    filtro_tipo = request.args.get('tipo', '')
    filtro_insersor = request.args.get('insersor', '').strip()
    filtro_pendente = request.args.get('pendente', '')

    # Apenas cursos que têm pelo menos uma disciplina
    from sqlalchemy import exists as sql_exists, or_
    q = Course.query.filter(
        sql_exists().where(Discipline.course_id == Course.id)
    )
    if filtro_tipo:
        q = q.filter_by(tipo=filtro_tipo)
    todos = q.order_by(Course.tipo, Course.nome).all()

    # Filtrar por busca (nome do curso OU nome da disciplina)
    if busca:
        busca_low = busca.lower()
        disc_cids = {r[0] for r in db.session.query(Discipline.course_id)
                     .filter(Discipline.nome.ilike(f'%{busca}%')).all()}
        todos = [c for c in todos if busca_low in c.nome.lower() or c.id in disc_cids]

    # Filtrar por insersor (campo pode ser comma-separated)
    if filtro_insersor:
        fi_low = filtro_insersor.lower()
        todos = [c for c in todos if c.insersor and
                 any(fi_low == p.strip().lower() for p in c.insersor.split(','))]

    def _parse_ch(val):
        if not val:
            return 0
        import re
        m = re.search(r'\d+', str(val))
        return int(m.group()) if m else 0

    course_data = []
    for c in todos:
        discs = Discipline.query.filter_by(course_id=c.id).order_by(Discipline.ordem).all()
        if busca and busca.lower() not in c.nome.lower():
            discs = [d for d in discs if busca.lower() in d.nome.lower()]
        total_ch = sum(_parse_ch(d.carga) for d in discs)
        ok_count = sum(1 for d in discs if d.plataforma_ok)
        course_data.append({'course': c, 'disciplines': discs, 'total_ch': total_ch, 'ok_count': ok_count})

    # Filtrar apenas pendentes (alguma disciplina não concluída)
    if filtro_pendente == '1':
        course_data = [item for item in course_data
                       if item['ok_count'] < len(item['disciplines'])]

    tipos_disponiveis = [r[0] for r in db.session.query(Course.tipo).join(
        Discipline, Discipline.course_id == Course.id).distinct().all()]

    # Lista de insersores para o filtro (expandindo comma-separated)
    ins_set = set()
    for c in Course.query.filter(
        sql_exists().where(Discipline.course_id == Course.id)
    ).all():
        if c.insersor:
            for p in c.insersor.split(','):
                p = p.strip()
                if p:
                    ins_set.add(p)
    insersores_disponiveis = sorted(ins_set)

    total_pendentes = sum(1 for item in course_data
                          if item['ok_count'] < len(item['disciplines']))

    # Pendentes por tipo (apenas cursos com disciplinas, sem filtro atual)
    todos_para_chips = Course.query.filter(
        sql_exists().where(Discipline.course_id == Course.id)
    ).all()
    pendentes_por_tipo = {}
    for c in todos_para_chips:
        discs_c = Discipline.query.filter_by(course_id=c.id).all()
        ok_c = sum(1 for d in discs_c if d.plataforma_ok)
        if ok_c < len(discs_c):
            pendentes_por_tipo[c.tipo] = pendentes_por_tipo.get(c.tipo, 0) + 1

    return render_template('matrizes.html', course_data=course_data, busca=busca,
                           filtro_tipo=filtro_tipo, tipos_disponiveis=tipos_disponiveis,
                           filtro_insersor=filtro_insersor, filtro_pendente=filtro_pendente,
                           insersores_disponiveis=insersores_disponiveis,
                           total_pendentes=total_pendentes,
                           pendentes_por_tipo=pendentes_por_tipo)

@app.route('/matrizes/marcar-tudo', methods=['POST'])
@login_required
def matrizes_marcar_tudo():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.json or {}
    marcar = data.get('marcar', True)
    filtro_tipo = data.get('tipo', '')
    filtro_status = data.get('status', '')
    filtro_insersor = data.get('insersor', '').strip()
    now = datetime.utcnow()
    q = Course.query
    if filtro_tipo:
        q = q.filter_by(tipo=filtro_tipo)
    if filtro_status:
        q = q.filter_by(status=filtro_status)
    cursos = q.all()
    if filtro_insersor:
        fi_low = filtro_insersor.lower()
        cursos = [c for c in cursos if c.insersor and
                  any(fi_low == p.strip().lower() for p in c.insersor.split(','))]
    course_ids = [c.id for c in cursos]
    discs = Discipline.query.filter(Discipline.course_id.in_(course_ids)).all()
    for d in discs:
        d.plataforma_ok = marcar
        d.plataforma_em = now if marcar else None
    db.session.commit()
    parts = []
    if filtro_tipo: parts.append(f'tipo={filtro_tipo}')
    if filtro_status: parts.append(f'status={filtro_status}')
    if filtro_insersor: parts.append(f'insersor={filtro_insersor}')
    filtro_desc = ' (' + ', '.join(parts) + ')' if parts else ' (todos)'
    log_action(session['user_id'], session['username'],
               'marcar_tudo' if marcar else 'desmarcar_tudo',
               'discipline', None,
               f'{"Marcou" if marcar else "Desmarcou"} {len(discs)} disciplinas{filtro_desc}')
    return jsonify({'ok': True, 'total': len(discs), 'marcar': marcar})

@app.route('/matrizes/relatorio')
@login_required
def matrizes_relatorio():
    filtro_tipo = request.args.get('tipo', '')
    filtro_status = request.args.get('status', '')
    from sqlalchemy import exists as sql_exists
    import re
    q = Course.query.filter(sql_exists().where(Discipline.course_id == Course.id))
    if filtro_tipo:
        q = q.filter_by(tipo=filtro_tipo)
    if filtro_status:
        q = q.filter_by(status=filtro_status)
    todos = q.order_by(Course.tipo, Course.nome).all()

    def _parse_ch(val):
        if not val: return 0
        m = re.search(r'\d+', str(val))
        return int(m.group()) if m else 0

    course_data = []
    for c in todos:
        discs = Discipline.query.filter_by(course_id=c.id).order_by(Discipline.ordem).all()
        total_ch = sum(_parse_ch(d.carga) for d in discs)
        course_data.append({'course': c, 'disciplines': discs, 'total_ch': total_ch})

    tipos_disponiveis = [r[0] for r in db.session.query(Course.tipo).join(
        Discipline, Discipline.course_id == Course.id).distinct().all()]
    status_list = ['ativo', 'em_edicao', 'finalizado', 'descontinuado', 'oculto']

    return render_template('matrizes_relatorio.html', course_data=course_data,
                           filtro_tipo=filtro_tipo, filtro_status=filtro_status,
                           tipos_disponiveis=tipos_disponiveis, status_list=status_list,
                           now=datetime.utcnow())

@app.route('/matrizes/exportar-excel')
@login_required
def matrizes_exportar_excel():
    import openpyxl, re
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from sqlalchemy import exists as sql_exists

    filtro_tipo   = request.args.get('tipo', '')
    filtro_status = request.args.get('status', '')

    q = Course.query.filter(sql_exists().where(Discipline.course_id == Course.id))
    if filtro_tipo:   q = q.filter_by(tipo=filtro_tipo)
    if filtro_status: q = q.filter_by(status=filtro_status)
    todos = q.order_by(Course.tipo, Course.nome).all()

    def _parse_ch(val):
        if not val: return 0
        m = re.search(r'\d+', str(val))
        return int(m.group()) if m else 0

    TIPO_LABELS = {
        'pos':'Pós-Graduação','profissionalizante':'Profissionalizante','rapido':'Rápido',
        'pacote':'Pacote','terceiros':'Terceiros','evento':'Evento',
        'pratica_conectada':'Prática Conectada','pratica_estagio':'Prática Estágio',
        'projeto_ambiental':'Proj. Ambiental','ggbr':'GGBR','integra_edu':'Integra Edu',
    }

    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo por curso ─────────────────────────────────────────────
    ws_res = wb.active
    ws_res.title = 'Resumo'

    hdr_fill   = PatternFill('solid', fgColor='6366F1')
    hdr_font   = Font(bold=True, color='FFFFFF', size=10)
    ok_fill    = PatternFill('solid', fgColor='DCFCE7')
    pend_fill  = PatternFill('solid', fgColor='FEF9C3')
    thin       = Side(style='thin', color='CBD5E1')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap       = Alignment(wrap_text=True, vertical='top')

    res_headers = ['Curso', 'Tipo', 'Área', 'Status', 'CH Total', 'Duração',
                   'Valor (R$)', 'Insersor', 'Cupom', 'Ano', 'Link de Venda',
                   'Total Disc.', 'Na Plataforma', 'Pendentes', '% Concluído']
    for col, h in enumerate(res_headers, 1):
        c = ws_res.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

    ws_res.row_dimensions[1].height = 30

    for row_idx, curso in enumerate(todos, 2):
        discs = Discipline.query.filter_by(course_id=curso.id).order_by(Discipline.ordem).all()
        total_ch  = sum(_parse_ch(d.carga) for d in discs)
        ok_count  = sum(1 for d in discs if d.plataforma_ok)
        pend      = len(discs) - ok_count
        pct       = round(ok_count / len(discs) * 100) if discs else 0
        values = [
            curso.nome,
            TIPO_LABELS.get(curso.tipo, curso.tipo),
            curso.area or '',
            curso.status.replace('_',' ').title(),
            f'{total_ch}h' if total_ch else (curso.horas or ''),
            curso.meses or '',
            curso.valor or '',
            curso.insersor or '',
            curso.cupom or '',
            curso.ano or '',
            curso.link_venda or '',
            len(discs), ok_count, pend,
            f'{pct}%',
        ]
        row_fill = ok_fill if pct == 100 else (pend_fill if pct > 0 else None)
        for col, val in enumerate(values, 1):
            cell = ws_res.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = wrap
            if row_fill and col <= 11:
                cell.fill = row_fill

    # Column widths resumo
    widths = [50, 18, 14, 14, 10, 10, 10, 20, 12, 8, 40, 10, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        ws_res.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws_res.freeze_panes = 'A2'

    # ── Aba 2: Matrizes completas ────────────────────────────────────────────
    ws_mat = wb.create_sheet('Matrizes')

    mat_headers = ['Curso', 'Tipo', 'Área', 'Status', 'Insersor',
                   'Módulo', '#', 'Disciplina', 'CH', 'Professor', 'Titulação',
                   'Plataforma', 'Data Inserção']
    for col, h in enumerate(mat_headers, 1):
        c = ws_mat.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border
    ws_mat.row_dimensions[1].height = 28

    row_idx = 2
    for curso in todos:
        discs = Discipline.query.filter_by(course_id=curso.id).order_by(Discipline.ordem).all()
        for d in discs:
            values = [
                curso.nome,
                TIPO_LABELS.get(curso.tipo, curso.tipo),
                curso.area or '',
                curso.status.replace('_',' ').title(),
                curso.insersor or '',
                d.modulo or '',
                d.ordem,
                d.nome,
                d.carga or '',
                d.professor or '',
                d.titulacao or '',
                'Sim' if d.plataforma_ok else 'Pendente',
                d.plataforma_em.strftime('%d/%m/%Y') if d.plataforma_ok and d.plataforma_em else '',
            ]
            row_fill = ok_fill if d.plataforma_ok else pend_fill
            for col, val in enumerate(values, 1):
                cell = ws_mat.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = wrap
                cell.fill = row_fill
            row_idx += 1

    # Column widths matrizes
    mat_widths = [45, 18, 12, 14, 18, 12, 5, 45, 8, 22, 18, 10, 12]
    for i, w in enumerate(mat_widths, 1):
        ws_mat.column_dimensions[get_column_letter(i)].width = w
    ws_mat.freeze_panes = 'A2'

    # ── Salva e envia ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime as dt
    ts = dt.now().strftime('%Y%m%d_%H%M')
    fname = f'matrizes_inova_{ts}.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

# ─── HISTÓRICO / AUDIT ─────────────────────────────────────────────────────────

@app.route('/historico')
@perm_check('can_view_historico')
def historico():
    page = request.args.get('page', 1, type=int)
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(page=page, per_page=50)
    return render_template('historico.html', logs=logs)

# ─── USUÁRIOS ──────────────────────────────────────────────────────────────────

@app.route('/usuarios')
@admin_required
def usuarios():
    users = User.query.order_by(User.username).all()
    # todos os cursos, sem exceção de tipo ou status
    todos_cursos = Course.query.all()

    stats = {}
    for u in users:
        meus = [c for c in todos_cursos if c.insersor and _insersor_contains(c.insersor, u.username)]
        por_status = {}
        for c in meus:
            por_status[c.status] = por_status.get(c.status, 0) + 1
        por_tipo = {}
        for c in meus:
            por_tipo[c.tipo] = por_tipo.get(c.tipo, 0) + 1
        stats[u.id] = {
            'total':      len(meus),
            'ativos':     por_status.get('ativo', 0),
            'em_edicao':  por_status.get('em_edicao', 0),
            'finalizado': por_status.get('finalizado', 0),
            'desc':       por_status.get('descontinuado', 0),
            'oculto':     por_status.get('oculto', 0),
            'por_tipo':   por_tipo,
        }

    TIPO_LABEL = {
        'pos': 'Pós', 'profissionalizante': 'Profis.', 'rapido': 'Rápido',
        'pacote': 'Pacote', 'terceiros': 'Terceiros', 'evento': 'Evento',
        'pratica_conectada': 'Prática', 'pratica_estagio': 'Estágio',
        'projeto_ambiental': 'Proj. Amb.', 'ggbr': 'GGBR', 'integra_edu': 'Integra',
    }
    return render_template('usuarios.html', users=users, stats=stats, tipo_label=TIPO_LABEL)


@app.route('/usuarios/<int:id>/cursos')
@admin_required
def usuario_cursos(id):
    u = User.query.get_or_404(id)
    todos_cursos = Course.query.filter(Course.insersor != None, Course.insersor != '')\
                               .order_by(Course.nome).all()
    meus = [c for c in todos_cursos if _insersor_contains(c.insersor, u.username)]

    disc_stats = {}
    for c in meus:
        total = Discipline.query.filter_by(course_id=c.id).count()
        pend  = Discipline.query.filter_by(course_id=c.id, plataforma_ok=False).count()
        disc_stats[c.id] = {'total': total, 'pend': pend, 'ok': total - pend}

    TIPO_LABEL = {
        'pos': 'Pós-Graduação', 'profissionalizante': 'Profissionalizante',
        'rapido': 'Rápido', 'pacote': 'Pacote', 'terceiros': 'Terceiros',
        'evento': 'Evento', 'pratica_conectada': 'Prática Conectada',
        'pratica_estagio': 'Prática Estágio', 'projeto_ambiental': 'Proj. Ambiental',
        'ggbr': 'GGBR', 'integra_edu': 'Integra Edu',
    }
    return render_template('usuario_cursos.html', u=u, cursos=meus,
                           disc_stats=disc_stats, tipo_label=TIPO_LABEL)

def _validar_email_institucional(email):
    email = (email or '').strip().lower()
    if not email or not email.endswith(EMAIL_DOMINIO_PERMITIDO):
        return None
    return email

@app.route('/usuarios/novo', methods=['GET','POST'])
@admin_required
def usuario_novo():
    if request.method == 'POST':
        d = request.form
        email = _validar_email_institucional(d.get('email'))
        if User.query.filter_by(username=d['username']).first():
            flash('Usuário já existe.', 'danger')
        elif not email:
            flash(f'O e-mail precisa ser institucional ({EMAIL_DOMINIO_PERMITIDO}).', 'danger')
        elif User.query.filter_by(email=email).first():
            flash('Já existe um usuário com esse e-mail.', 'danger')
        else:
            perms = _perms_from_form(request.form)
            u = User(username=d['username'], nome=d.get('nome', '').strip(), email=email,
                     password=hash_pw(d['password']),
                     role=d['role'], permissoes=json.dumps(perms), must_change_password=True)
            db.session.add(u)
            db.session.commit()
            log_action(session['user_id'], session['username'], 'criar', 'user', u.id, u.username)
            flash('Usuário criado!', 'success')
            return redirect(url_for('usuarios'))
    return render_template('usuario_form.html', user=None)

@app.route('/usuarios/<int:id>/editar', methods=['GET','POST'])
@admin_required
def usuario_editar(id):
    u = User.query.get_or_404(id)
    if request.method == 'POST':
        d = request.form
        novo_username = d.get('username', '').strip()
        if novo_username and novo_username != u.username:
            existente = User.query.filter_by(username=novo_username).first()
            if existente:
                flash('Já existe um usuário com esse nome.', 'danger')
                return render_template('usuario_form.html', user=u)
            u.username = novo_username
        novo_email = _validar_email_institucional(d.get('email'))
        if not novo_email:
            flash(f'O e-mail precisa ser institucional ({EMAIL_DOMINIO_PERMITIDO}).', 'danger')
            return render_template('usuario_form.html', user=u)
        if novo_email != u.email and User.query.filter_by(email=novo_email).first():
            flash('Já existe um usuário com esse e-mail.', 'danger')
            return render_template('usuario_form.html', user=u)
        u.email = novo_email
        u.nome = d.get('nome', '').strip()
        u.role = d['role']
        u.permissoes = json.dumps(_perms_from_form(d))
        if d.get('password'):
            u.password = hash_pw(d['password'])
            u.must_change_password = True
        db.session.commit()
        log_action(session['user_id'], session['username'], 'editar', 'user', id, u.username)
        flash('Usuário atualizado!', 'success')
        return redirect(url_for('usuarios'))
    return render_template('usuario_form.html', user=u)

def _perms_from_form(d):
    keys = [
        'cursos_editar', 'cursos_excluir',
        'cupons_gerenciar', 'reembolsos_gerenciar',
        'historico_ver', 'usuarios_gerenciar', 'backup_gerenciar',
        'block_cupons', 'block_reembolsos', 'block_historico', 'block_trocar_senha',
    ]
    return {k: (d.get(f'perm_{k}') == 'on') for k in keys}

@app.route('/usuarios/<int:id>/excluir', methods=['POST'])
@admin_required
def usuario_excluir(id):
    u = User.query.get_or_404(id)
    if u.id == session['user_id']:
        flash('Você não pode excluir sua própria conta.', 'danger')
        return redirect(url_for('usuarios'))
    username = u.username
    # Cursos e registros de histórico guardam o nome em texto separadamente
    # (Course.insersor, AuditLog.username), então desvincular o ID aqui não perde
    # o histórico — só evita a violação de chave estrangeira ao excluir o usuário.
    Course.query.filter_by(created_by=u.id).update({'created_by': None})
    AuditLog.query.filter_by(user_id=u.id).update({'user_id': None})
    db.session.delete(u)
    db.session.commit()
    log_action(session['user_id'], session['username'], 'excluir', 'user', id, username)
    flash(f'Usuário "{username}" excluído.', 'success')
    return redirect(url_for('usuarios'))

# ─── BACKUP ────────────────────────────────────────────────────────────────────

@app.route('/backup/manual', methods=['POST'])
@admin_required
def backup_manual():
    if not app.config.get('SQLALCHEMY_DATABASE_URI', '').startswith('sqlite'):
        flash('Backup de arquivo está disponível apenas no modo local. Os dados no Supabase são gerenciados na nuvem.', 'info')
        return redirect(url_for('dashboard'))
    with app.app_context():
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        src = os.path.join(app.instance_path, 'inova.db')
        fname = f'manual_{ts}.zip'
        fpath = os.path.join('backups', fname)
        os.makedirs('backups', exist_ok=True)
        with zipfile.ZipFile(fpath, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.write(src, 'inova.db')
        size_kb = os.path.getsize(fpath) / 1024
        rec = BackupRecord(filename=fname, size_kb=round(size_kb,2), tipo='manual')
        db.session.add(rec)
        db.session.commit()
    flash('Backup manual criado com sucesso!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/backup/download/<int:id>')
@admin_required
def backup_download(id):
    rec = BackupRecord.query.get_or_404(id)
    fpath = os.path.join('backups', rec.filename)
    if not os.path.exists(fpath):
        flash('Arquivo de backup não disponível neste ambiente.', 'warning')
        return redirect(url_for('backup_lista'))
    return send_file(fpath, as_attachment=True, download_name=rec.filename)

@app.route('/backup/lista')
@admin_required
def backup_lista():
    bks = BackupRecord.query.order_by(BackupRecord.created_at.desc()).all()
    return render_template('backups.html', bks=bks)

# ─── SEED DATA ─────────────────────────────────────────────────────────────────

def _import_excel():
    import re as _re2

    def is_numero(val):
        try: return int(str(val).strip()) > 0
        except: return False

    def limpar_horas(val):
        if val is None: return ''
        m = _re2.match(r'^(\d+\.?\d*)', str(val).strip())
        return m.group(1) if m else ''

    def limpar_valor(val):
        if val is None: return ''
        s = str(val).strip()
        return s if s not in ('-', '') else ''

    def pd(v):
        return v.date() if isinstance(v, datetime) else None

    admin = User.query.filter_by(username='admin').first()
    admin_id = admin.id if admin else None

    def ac(nome, tipo, area, horas, valor, link, link_img='', desc='', obs='', status='ativo', insersor='', meses='', cupom='', dono=''):
        if not nome or str(nome).strip() == '': return None
        c = Course(nome=str(nome).strip()[:300], tipo=tipo,
                   area=str(area or '').strip()[:100], horas=str(horas or '').strip()[:20],
                   meses=str(meses or '').strip()[:20], valor=str(valor or '').strip()[:50],
                   link_venda=str(link or '').strip(), link_imagem=str(link_img or '').strip(),
                   descricao=str(desc or '').strip(), obs=str(obs or '').strip(),
                   status=status, insersor=str(insersor or '').strip()[:100],
                   cupom=str(cupom or '').strip()[:100], dono=str(dono or '').strip(),
                   created_by=admin_id)
        db.session.add(c)
        return c

    try:
        import openpyxl
        excel_path = os.path.join(os.path.dirname(__file__), 'CURSOS INOVA - LINKS (1).xlsx')
        wb = openpyxl.load_workbook(excel_path)

        # PÓS-GRADUAÇÃO — só linhas onde col[0] é número (ignora linhas de disciplinas)
        for shname in wb.sheetnames:
            if 'MATRIZES' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not is_numero(row[0]): continue
                    area_val = str(row[2] or '').strip()
                    if area_val.endswith('h') and area_val[:-1].isdigit(): continue
                    ac(row[1], 'pos', row[2], limpar_horas(row[3]), limpar_valor(row[5]),
                       str(row[7] or ''), obs=str(row[8] or ''), insersor=str(row[6] or ''),
                       meses=str(row[4] or ''))
                break

        # PROFISSIONALIZANTES
        for shname in wb.sheetnames:
            if 'PROFISSIONALIZANTE' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    ac(row[1], 'profissionalizante', row[2], limpar_horas(row[3]), '',
                       str(row[4] or ''), link_img=str(row[6] or ''),
                       desc=str(row[5] or ''), obs=str(row[7] or ''), insersor='INOVA')
                break

        # RÁPIDOS
        for shname in wb.sheetnames:
            if 'RÁPIDOS INOVA' in shname or 'RAPIDOS INOVA' in shname:
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    ac(row[1], 'rapido', row[2], limpar_horas(row[3]), limpar_valor(row[5]),
                       str(row[4] or ''), link_img=str(row[7] or ''),
                       desc=str(row[6] or ''), obs=str(row[8] or ''), insersor='INOVA')
                break

        # PACOTES
        for shname in wb.sheetnames:
            if shname.upper() == 'PACOTE CURSOS':
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    ac(row[1], 'pacote', row[2], limpar_horas(row[3]), limpar_valor(row[8]),
                       str(row[4] or ''), obs=str(row[6] or ''), insersor=str(row[5] or ''))
                break

        # TERCEIROS
        for shname in wb.sheetnames:
            if 'TERCEIROS' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    obs = str(row[10] or '').strip()
                    s = 'descontinuado' if 'DESCONTINUADO' in obs.upper() else 'ativo'
                    ac(row[1], 'terceiros', row[5], limpar_horas(row[3]), limpar_valor(row[2]),
                       str(row[6] or ''), desc=str(row[8] or ''), obs=obs, dono=str(row[7] or ''))
                break

        # EVENTOS
        for shname in wb.sheetnames:
            if shname.upper() == 'EVENTOS':
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[2]: continue
                    obs_ev = f"Tipo: {row[1] or ''} | Data: {row[5] or ''}" + (f" | {row[8]}" if row[8] else "")
                    ac(row[2], 'evento', 'EVENTO', limpar_horas(row[3]), limpar_valor(row[6]),
                       str(row[4] or ''), obs=obs_ev, insersor='INOVA')
                break

        # PRÁTICAS CONECTADAS
        for shname in wb.sheetnames:
            if 'CONECTADA' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[1]: continue
                    ac(row[1], 'pratica_conectada', str(row[0] or ''), '', '',
                       str(row[3] or ''), status='oculto', insersor='INOVA', cupom=str(row[2] or ''))
                break

        # PRÁTICAS ESTÁGIO
        for shname in wb.sheetnames:
            if 'ESTAGIO' in shname.upper() or 'ESTÁGIO' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    ac(row[0], 'pratica_estagio', str(row[1] or ''), '', '',
                       str(row[5] or ''), insersor=str(row[3] or ''), cupom=str(row[4] or ''))
                break

        # PROJ. AMBIENTES PROF
        for shname in wb.sheetnames:
            if 'AMBIENTES' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    ac(row[1], 'projeto_ambiental', '', '', '', str(row[3] or ''),
                       status='oculto', insersor='INOVA', cupom=str(row[2] or ''), obs=str(row[4] or ''))
                break

        # GGBR
        for shname in wb.sheetnames:
            if 'GGBR' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not is_numero(row[0]): continue
                    ac(row[1], 'ggbr', str(row[2] or ''), limpar_horas(row[3]),
                       limpar_valor(row[5]), str(row[4] or ''), insersor='INOVA')
                break

        # INTEGRA EDU
        for shname in wb.sheetnames:
            if 'INTEGRA' in shname.upper():
                ws = wb[shname]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    ac(row[0], 'integra_edu', '', limpar_horas(row[1]), '',
                       str(row[2] or ''), obs=str(row[3] or ''), insersor='INOVA')
                break

        db.session.commit()

        # CUPONS — só importa se não há cupons ainda
        if Coupon.query.count() == 0:
            for shname in wb.sheetnames:
                if shname.upper() == 'CUPOM':
                    ws = wb[shname]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0] or str(row[0]).strip() in ('NOME', ''): continue
                        try:
                            cp = Coupon(nome=str(row[0]).strip()[:100], quantidade=int(row[1] or 0),
                                       desconto=float(row[2] or 0), cursos_tipo=str(row[3] or '').strip(),
                                       limite_curso=int(row[4] or 1), uso_unico=(str(row[5] or '').upper()=='SIM'),
                                       data_inicial=pd(row[6]), data_final=pd(row[7]) if len(row) > 7 else None,
                                       obs=str(row[8] or '').strip() if len(row) > 8 else '')
                            db.session.add(cp)
                        except: pass
                    break

        # REEMBOLSOS — só importa se não há reembolsos ainda
        if Refund.query.count() == 0:
            for shname in wb.sheetnames:
                if 'REEMBOLSO' in shname.upper():
                    ws = wb[shname]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[1] or str(row[1]).strip() in ('NOME ALUNO', ''): continue
                        try:
                            r = Refund(colab=str(row[0] or '').strip(), nome_aluno=str(row[1] or '').strip(),
                                      data_compra=pd(row[2]), data_solicitacao=pd(row[3]),
                                      valor=float(str(row[4] or 0).replace(',', '.') or 0),
                                      valor_estorno=float(str(row[5] or 0).replace(',', '.') or 0),
                                      nome_curso=str(row[6] or '').strip(), categoria=str(row[7] or '').strip(),
                                      data_aprovacao=pd(row[10]) if len(row) > 10 else None,
                                      motivo=str(row[11] or '').strip() if len(row) > 11 else '')
                            db.session.add(r)
                        except: pass
                    break

        db.session.commit()

        # ── DISCIPLINAS DAS MATRIZES (PÓS) ────────────────────────────────
        import unicodedata as _ud

        def norm_nome(s):
            s = _re2.sub(r'\s+', ' ', str(s).upper().strip())
            s = ''.join(c for c in _ud.normalize('NFKD', s) if not _ud.combining(c))
            return _re2.sub(r'\s+', ' ', s).strip()

        def norm_compact(s):
            return _re2.sub(r'\s+', '', norm_nome(s))

        def nome_sim(a, b):
            na, nb = norm_nome(a), norm_nome(b)
            if na == nb:
                return True
            # Compara sem espaços (resolve \xa0 embutido)
            ca, cb = norm_compact(a), norm_compact(b)
            if ca == cb:
                return True
            # Prefixo longo (evita colisão entre "EDUCACAO DO FUTURO" e "EDUCACAO INTEGRADA...")
            for n in [55, 50, 45, 40]:
                if len(ca) >= n and len(cb) >= n and ca[:n] == cb[:n]:
                    return True
            # Prefixo com espaços (normalizado) — só acima de 35 chars
            for n in [45, 40, 35]:
                if len(na) >= n and len(nb) >= n and na[:n] == nb[:n]:
                    return True
            return False

        pos_courses = Course.query.filter_by(tipo='pos').all()

        sheet_m = next((s for s in wb.sheetnames if 'MATRIZES' in s.upper()), None)
        if sheet_m and pos_courses:
            ws_m = wb[sheet_m]
            all_rows = list(ws_m.iter_rows(min_row=3, values_only=True))

            # Encontrar onde terminam as linhas de cursos e começam as matrizes
            matrix_start = len(all_rows)
            for i, row in enumerate(all_rows):
                try: int(str(row[0] or '').strip())
                except:
                    matrix_start = i
                    break

            cur_id = None
            in_disc = False
            disc_ordem = 0
            cur_modulo = None

            for row in all_rows[matrix_start:]:
                c0 = str(row[0] or '').strip()
                c1 = str(row[1] or '').strip()
                c2 = str(row[2] or '').strip()
                c3 = str(row[3] or '').strip()

                if not c0 and not c1 and not c2 and not c3:
                    in_disc = False
                    continue

                if c1.upper() == 'DISCIPLINAS':
                    in_disc = True
                    disc_ordem = 0
                    cur_modulo = None
                    continue

                if not c0 and not c1 and c2:
                    continue

                if not c2:
                    cname = c0 if (c0 and not c1) else c1
                    if cname and cname.upper() not in ('PROFESSOR', 'PROFESSORES', 'DISCIPLINAS'):
                        in_disc = False
                        disc_ordem = 0
                        cur_modulo = None
                        cur_id = None
                        match = next((c for c in pos_courses if nome_sim(c.nome, cname)), None)
                        if match:
                            existing = Discipline.query.filter_by(course_id=match.id).count()
                            cur_id = match.id if existing == 0 else None
                    continue

                if in_disc and c1 and c2 and cur_id:
                    if c0 and not c0.isdigit():
                        cur_modulo = c0
                        disc_ordem += 1
                    elif c0 and c0.isdigit():
                        disc_ordem = int(c0)
                    else:
                        disc_ordem += 1
                    db.session.add(Discipline(
                        course_id=cur_id, modulo=cur_modulo,
                        ordem=disc_ordem, nome=c1, carga=c2 or None,
                        professor=c3 or None
                    ))

            db.session.commit()

        # ── DISCIPLINAS DOS PROFISSIONALIZANTES ───────────────────────────
        # Layout: 2 matrizes lado a lado por bloco de linhas
        # Bloco 1: C9=modulo, C10=ordem, C11=nome, C12=ch
        # Bloco 2: C14=modulo, C15=ordem, C16=nome, C17=ch
        # O nome do curso fica na linha ANTES do cabeçalho "DISCIPLINAS"
        prof_courses = Course.query.filter_by(tipo='profissionalizante').all()
        sheet_prof = next((s for s in wb.sheetnames if 'PROFISSIONALIZANTE' in s.upper()), None)
        if sheet_prof and prof_courses:
            prof_map = {}
            prof_map_compact = {}
            for c in prof_courses:
                prof_map[norm_nome(c.nome)] = c
                prof_map_compact[norm_compact(c.nome)] = c

            def _match_prof(nome_excel):
                key = norm_nome(nome_excel)
                if key in prof_map:
                    return prof_map[key]
                ck = norm_compact(nome_excel)
                if ck in prof_map_compact:
                    return prof_map_compact[ck]
                for n in [40, 35, 30, 25, 20]:
                    for k, c in prof_map_compact.items():
                        if len(ck) >= n and len(k) >= n and ck[:n] == k[:n]:
                            return c
                return None

            ws_p = wb[sheet_prof]
            rows_p = list(ws_p.iter_rows(min_row=1, values_only=True))
            prev_row = None
            cur1_id = None
            cur2_id = None
            mod1 = None
            mod2 = None

            def _cell(row, idx):
                return str(row[idx] or '').strip() if row and idx < len(row) else ''

            for row in rows_p:
                c11 = _cell(row, 11)
                c16 = _cell(row, 16)

                # Linha de cabeçalho: detecta "DISCIPLINAS" em C11 ou C16
                if c11.upper() == 'DISCIPLINAS' or c16.upper() == 'DISCIPLINAS':
                    if prev_row is not None:
                        nome1 = _cell(prev_row, 9)
                        nome2 = _cell(prev_row, 14)
                        m1 = _match_prof(nome1) if nome1 else None
                        m2 = _match_prof(nome2) if nome2 else None
                        if m1 and Discipline.query.filter_by(course_id=m1.id).count() == 0:
                            cur1_id = m1.id
                        else:
                            cur1_id = None
                        if m2 and Discipline.query.filter_by(course_id=m2.id).count() == 0:
                            cur2_id = m2.id
                        else:
                            cur2_id = None
                    mod1 = None
                    mod2 = None
                    prev_row = row
                    continue

                # Bloco 1: C10=ordem (número), C11=nome disciplina, C12=carga
                c9  = _cell(row, 9)
                c10 = _cell(row, 10)
                c12 = _cell(row, 12)
                if c9.startswith('Mód'):
                    mod1 = c9
                if c10.isdigit() and c11 and c11.upper() not in ('DISCIPLINAS', 'CH') and cur1_id:
                    db.session.add(Discipline(
                        course_id=cur1_id, modulo=mod1,
                        ordem=int(c10), nome=c11, carga=c12 or None
                    ))

                # Bloco 2: C15=ordem (número), C16=nome disciplina, C17=carga
                c14 = _cell(row, 14)
                c15 = _cell(row, 15)
                c17 = _cell(row, 17)
                if c14.startswith('Mód'):
                    mod2 = c14
                if c15.isdigit() and c16 and c16.upper() not in ('DISCIPLINAS', 'CH') and cur2_id:
                    db.session.add(Discipline(
                        course_id=cur2_id, modulo=mod2,
                        ordem=int(c15), nome=c16, carga=c17 or None
                    ))

                prev_row = row

            db.session.commit()

        print("[OK] Dados importados com sucesso!")
    except FileNotFoundError:
        print("[INFO] Arquivo Excel nao encontrado.")
        db.session.rollback()
    except Exception as e:
        print(f"[ERRO] Ao importar Excel: {e}")
        db.session.rollback()

def seed_data():
    if User.query.count() == 0:
        # Senhas padrão temporárias — troca obrigatória no primeiro acesso.
        admin = User(username='admin', email='admin' + EMAIL_DOMINIO_PERMITIDO,
                     password=hash_pw('inova2024'), role='admin', must_change_password=True)
        junior = User(username='junior', email='junior' + EMAIL_DOMINIO_PERMITIDO,
                      password=hash_pw('inova2024'), role='editor', must_change_password=True)
        felipe = User(username='felipe', email='felipe' + EMAIL_DOMINIO_PERMITIDO,
                      password=hash_pw('inova2024'), role='editor', must_change_password=True)
        viewer = User(username='visualizador', email='visualizador' + EMAIL_DOMINIO_PERMITIDO,
                      password=hash_pw('inova2024'), role='viewer', must_change_password=True)
        db.session.add_all([admin, junior, felipe, viewer])
        db.session.commit()
    if Course.query.count() == 0:
        _import_excel()

@app.route('/admin/reimportar', methods=['GET', 'POST'])
@admin_required
def admin_reimportar():
    # Remove apenas cursos, disciplinas e cupons — preserva reembolsos (dados reais)
    Discipline.query.delete()
    Course.query.delete()
    Coupon.query.delete()
    db.session.commit()
    _import_excel()
    flash('Dados limpos e reimportados com sucesso! Reembolsos foram preservados.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/importar-disciplinas', methods=['POST'])
@admin_required
def admin_importar_disciplinas():
    """Importa apenas disciplinas do Excel sem apagar dados existentes."""
    try:
        total = _importar_so_disciplinas()
        flash(f'{total} disciplina(s) importada(s) com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao importar disciplinas: {e}', 'danger')
    return redirect(url_for('dashboard'))

def _importar_so_disciplinas():
    import re as _re2, unicodedata as _ud2, openpyxl as _opx
    excel_path = os.path.join(os.path.dirname(__file__), 'CURSOS INOVA - LINKS (1).xlsx')
    if not os.path.exists(excel_path):
        raise FileNotFoundError('Arquivo Excel não encontrado.')
    wb = _opx.load_workbook(excel_path)

    def _norm(s):
        s = _re2.sub(r'\s+', ' ', str(s).upper().strip())
        return _re2.sub(r'\s+', ' ',
               ''.join(c for c in _ud2.normalize('NFKD', s) if not _ud2.combining(c))).strip()
    def _nc(s): return _re2.sub(r'\s+', '', _norm(s))
    def _sim(a, b):
        na, nb = _norm(a), _norm(b)
        if na == nb: return True
        ca, cb = _nc(a), _nc(b)
        if ca == cb: return True
        for n in [55, 50, 45, 40]:
            if len(ca) >= n and len(cb) >= n and ca[:n] == cb[:n]: return True
        for n in [45, 40, 35]:
            if len(na) >= n and len(nb) >= n and na[:n] == nb[:n]: return True
        return False
    def _cell(row, idx): return str(row[idx] or '').strip() if row and idx < len(row) else ''

    tit_lookup = {}
    if 'TITULAÇÕES PÓS' in wb.sheetnames:
        for row in wb['TITULAÇÕES PÓS'].iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) > 4 and row[4]:
                tit_lookup[str(row[0]).strip().upper()] = str(row[4]).strip()

    total = 0

    # PÓS
    pos_courses = Course.query.filter_by(tipo='pos').all()
    sheet_m = next((s for s in wb.sheetnames if 'MATRIZES' in s.upper()), None)
    if sheet_m and pos_courses:
        ws = wb[sheet_m]
        all_rows = list(ws.iter_rows(min_row=3, values_only=True))
        mstart = next((i for i, r in enumerate(all_rows)
                       if not str(r[0] or '').strip().isdigit() and str(r[0] or '').strip()), len(all_rows))
        cur_id = None; in_disc = False; disc_ordem = 0; cur_mod = None
        for row in all_rows[mstart:]:
            c0,c1,c2,c3 = (_cell(row,i) for i in range(4))
            if not c0 and not c1 and not c2: in_disc = False; continue
            if c1.upper() == 'DISCIPLINAS': in_disc=True; disc_ordem=0; cur_mod=None; continue
            if not c0 and not c1 and c2: continue
            if not c2:
                cn = c0 if (c0 and not c1) else c1
                if cn and cn.upper() not in ('PROFESSOR','PROFESSORES','DISCIPLINAS'):
                    in_disc=False; disc_ordem=0; cur_mod=None; cur_id=None
                    m = next((c for c in pos_courses if _sim(c.nome, cn)), None)
                    if m and Discipline.query.filter_by(course_id=m.id).count()==0:
                        cur_id = m.id
                continue
            if in_disc and c1 and c2 and cur_id:
                if c0 and not c0.isdigit(): cur_mod=c0; disc_ordem+=1
                elif c0 and c0.isdigit(): disc_ordem=int(c0)
                else: disc_ordem+=1
                db.session.add(Discipline(course_id=cur_id, modulo=cur_mod,
                    ordem=disc_ordem, nome=c1, carga=c2 or None, professor=c3 or None,
                    titulacao=tit_lookup.get(c1.upper().strip(), '') or None))
                total += 1
        db.session.commit()

    # PROFISSIONALIZANTES
    prof_courses = Course.query.filter_by(tipo='profissionalizante').all()
    prof_map = {_norm(c.nome): c for c in prof_courses}
    prof_cmap = {_nc(c.nome): c for c in prof_courses}
    def _mp(n):
        if not n: return None
        if _norm(n) in prof_map: return prof_map[_norm(n)]
        if _nc(n) in prof_cmap: return prof_cmap[_nc(n)]
        for sz in [40,35,30,25,20]:
            for k,c in prof_cmap.items():
                if len(_nc(n))>=sz and len(k)>=sz and _nc(n)[:sz]==k[:sz]: return c
        return None
    sheet_p = next((s for s in wb.sheetnames if 'PROFISSIONALIZANTE' in s.upper()), None)
    if sheet_p and prof_courses:
        rows_p = list(wb[sheet_p].iter_rows(min_row=1, values_only=True))
        prev = None; c1id=c2id=None; m1=m2=None
        for row in rows_p:
            c11,c16 = _cell(row,11), _cell(row,16)
            if c11.upper()=='DISCIPLINAS' or c16.upper()=='DISCIPLINAS':
                if prev:
                    r1=_mp(_cell(prev,9)); r2=_mp(_cell(prev,14))
                    c1id = r1.id if r1 and Discipline.query.filter_by(course_id=r1.id).count()==0 else None
                    c2id = r2.id if r2 and Discipline.query.filter_by(course_id=r2.id).count()==0 else None
                m1=m2=None; prev=row; continue
            c9,c10,c12 = _cell(row,9),_cell(row,10),_cell(row,12)
            if c9.startswith('Mód'): m1=c9
            if c10.isdigit() and c11 and c11.upper() not in ('DISCIPLINAS','CH') and c1id:
                db.session.add(Discipline(course_id=c1id,modulo=m1,ordem=int(c10),nome=c11,carga=c12 or None)); total+=1
            c14,c15,c17 = _cell(row,14),_cell(row,15),_cell(row,17)
            if c14.startswith('Mód'): m2=c14
            if c15.isdigit() and c16 and c16.upper() not in ('DISCIPLINAS','CH') and c2id:
                db.session.add(Discipline(course_id=c2id,modulo=m2,ordem=int(c15),nome=c16,carga=c17 or None)); total+=1
            prev=row
        db.session.commit()

    return total

@app.route('/admin/excel-debug')
@admin_required
def admin_excel_debug():
    import openpyxl, os
    excel_path = os.path.join(os.path.dirname(__file__), 'CURSOS INOVA - LINKS (1).xlsx')
    wb = openpyxl.load_workbook(excel_path)
    linhas = [f"<b>Abas:</b> {', '.join(wb.sheetnames)}<br><br>"]

    # Todas as linhas não-vazias do Profissionalizantes após linha 43
    for shname in wb.sheetnames:
        if 'PROFISSIONALIZANTE' in shname.upper():
            ws = wb[shname]
            total = ws.max_row
            linhas.append(f"<b>Aba: {shname} — total de linhas: {total}</b><br>")
            # Mostrar primeiras linhas não-vazias após linha 43
            count = 0
            for i, row in enumerate(ws.iter_rows(min_row=44, values_only=True)):
                if any(c for c in row[:8] if c is not None and str(c).strip()):
                    cols = [str(c or '')[:40] for c in row[:8]]
                    linhas.append(f"Linha {i+44}: {' | '.join(cols)}<br>")
                    count += 1
                    if count >= 40:
                        break
            if count == 0:
                linhas.append("Nenhuma linha não-vazia encontrada após linha 43.<br>")
            break

    linhas.append("<br>")

    # Aba DISCIPLINAS PÓS
    if 'DISCIPLINAS PÓS' in wb.sheetnames:
        linhas.append("<b>Aba: DISCIPLINAS PÓS (primeiras 20 linhas)</b><br>")
        ws = wb['DISCIPLINAS PÓS']
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
            cols = [str(c or '')[:35] for c in row[:8]]
            linhas.append(f"Linha {i+1}: {' | '.join(cols)}<br>")

    return ''.join(linhas)

@app.route('/admin/status')
@admin_required
def admin_status():
    from sqlalchemy import func
    linhas = [f"<b>Status do banco:</b><br>"]
    linhas.append(f"Cursos: {Course.query.count()}<br>")
    linhas.append(f"Disciplinas: {Discipline.query.count()}<br>")
    linhas.append(f"Cupons: {Coupon.query.count()}<br>")
    linhas.append(f"Reembolsos: {Refund.query.count()}<br><br>")
    linhas.append("<b>Cursos por tipo:</b><br>")
    for tipo, qtd in db.session.query(Course.tipo, func.count(Course.id)).group_by(Course.tipo).all():
        com_disc = db.session.query(func.count(Course.id)).filter(
            Course.tipo == tipo,
            db.session.query(Discipline).filter(Discipline.course_id == Course.id).exists()
        ).scalar()
        linhas.append(f"{tipo}: {qtd} cursos, {com_disc} com disciplinas<br>")
    return ''.join(linhas)

# ─── INIT ──────────────────────────────────────────────────────────────────────

_db_ready = False

def _run_migrations():
    """Adiciona colunas que podem estar faltando em bancos mais antigos."""
    is_pg = _db_url.startswith('postgresql://')
    # Para PostgreSQL usa IF NOT EXISTS; para SQLite captura exceção
    migrations = [
        ("refund",     "concluido_manual", "BOOLEAN DEFAULT false"),
        ("refund",     "cpf",              "VARCHAR(20)"),
        ("refund",     "celular",          "VARCHAR(30)"),
        ("refund",     "pix",              "VARCHAR(200)"),
        ("refund",     "email_destino",    "VARCHAR(200)"),
        ("course",     "dono",             "TEXT"),
        ("course",     "ano",              "VARCHAR(10)"),
        ("course",     "extra_data",       "TEXT"),
        ("discipline", "cod_moodle",       "VARCHAR(50)"),
        ("discipline", "titulacao",        "VARCHAR(50)"),
        ("discipline", "plataforma_ok",    "BOOLEAN DEFAULT false"),
        ("discipline", "plataforma_em",    "TIMESTAMP"),
    ]
    with db.engine.connect() as conn:
        for table, col, dtype in migrations:
            try:
                if is_pg:
                    sql = f'ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {dtype}'
                else:
                    sql = f'ALTER TABLE {table} ADD COLUMN {col} {dtype}'
                conn.execute(db.text(sql))
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
        # tabela "user" precisa de aspas pois é palavra reservada em alguns DBs
        for col, dtype in [("permissoes", "TEXT DEFAULT '{}' "), ("email", "VARCHAR(200)"),
                           ("must_change_password", "BOOLEAN DEFAULT false"),
                           ("nome", "VARCHAR(200)")]:
            try:
                tbl = '"user"' if is_pg else 'user'
                sql = f'ALTER TABLE {tbl} ADD COLUMN {col} {dtype}'
                if is_pg:
                    sql = f'ALTER TABLE "user" ADD COLUMN IF NOT EXISTS {col} {dtype}'
                conn.execute(db.text(sql))
                conn.commit()
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass


@app.before_request
def ensure_db():
    global _db_ready
    if not _db_ready:
        try:
            db.create_all()
            _run_migrations()
            seed_data()
            _db_ready = True
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"<pre>Erro ao inicializar banco:\n{traceback.format_exc()}</pre>", 500

ROTAS_LIVRES_TROCA_SENHA = {'minha_conta', 'logout', 'login', 'static', 'esqueci_senha', 'resetar_senha'}

@app.before_request
def exigir_troca_senha():
    if request.endpoint in ROTAS_LIVRES_TROCA_SENHA or request.endpoint is None:
        return
    if 'user_id' not in session:
        return
    u = User.query.get(session['user_id'])
    if u and u.must_change_password and u.can_change_own_password():
        flash('Por segurança, troque sua senha antes de continuar.', 'danger')
        return redirect(url_for('minha_conta'))

if __name__ == '__main__':
    t = threading.Thread(target=backup_scheduler, daemon=True)
    t.start()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
